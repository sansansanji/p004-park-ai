# -*- coding: utf-8 -*-
"""
管理后台 API 服务
提供 RESTful 接口给前端管理界面
支持演示模式（无需 MySQL，使用内存模拟数据）

可以两种方式运行：
  1. python admin_api.py  ← 独立运行（本地调试用）
  2. 被 app.py 以 Blueprint 方式导入（生产推荐）
"""

import os
import sys
import json
import hashlib
import functools
import random
import traceback
from datetime import datetime, timedelta, date

from flask import Flask, Blueprint, request, jsonify, session, send_from_directory

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
CONFIG_DIR = os.path.join(BASE_DIR, 'config')
sys.path.insert(0, SCRIPT_DIR)

# 尝试连接数据库，失败则进入演示模式
DEMO_MODE = False
try:
    from db_manager import get_db, MemberDAO, MerchantDAO, ActivityDAO, CouponDAO, ArticleDAO, ChatLogDAO, ConfigDAO, ReplyRuleDAO, TagRuleDAO
    db_test = get_db()
    db_test.query_one("SELECT 1")
    print("=" * 50)
    print("[管理后台] [OK] 正式模式 - 已连接 MySQL 数据库")
    print("=" * 50)
except Exception as _db_err:
    DEMO_MODE = True
    print("=" * 50)
    print(f"[管理后台] [DEMO] 演示模式 - 数据库连接失败: {_db_err}")
    print("=" * 50)

# Blueprint（供 app.py 导入）
admin_bp = Blueprint('admin', __name__)

# 独立运行时用的 Flask app
app = Flask(__name__, static_folder=os.path.join(BASE_DIR, 'admin_frontend'))
app.secret_key = os.environ.get('SECRET_KEY', 'park-ai-admin-secret-2026')


# =============================================
# 演示模式数据持久化（重启不丢数据）
# =============================================
DEMO_DATA_FILE = os.path.join(BASE_DIR, 'data', 'demo_data.json')

def _load_demo_data():
    """从 JSON 文件恢复演示数据（如果有的话）"""
    if os.path.exists(DEMO_DATA_FILE):
        try:
            with open(DEMO_DATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return None

def _save_demo_data():
    """将演示数据保存到 JSON 文件"""
    try:
        os.makedirs(os.path.dirname(DEMO_DATA_FILE), exist_ok=True)
        with open(DEMO_DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'members': DEMO_MEMBERS,
                'merchants': DEMO_MERCHANTS,
                'activities': DEMO_ACTIVITIES,
                'coupons': DEMO_COUPONS,
                'articles': DEMO_ARTICLES,
            }, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[演示模式] 保存数据失败: {e}")


# =============================================
# 工具函数
# =============================================

def _fix_datetime(val):
    """将 ISO 格式日期转为 MySQL 兼容的 YYYY-MM-DD HH:MM:SS 格式"""
    if not val or not isinstance(val, str):
        return val
    # 处理 2026-05-03T16:00:00.000Z 格式
    val = val.replace('Z', '').replace('T', ' ')
    # 处理毫秒 2026-05-03 16:00:00.000
    if '.' in val:
        val = val.split('.')[0]
    return val


def _safe_execute(dao_method, *args, **kwargs):
    """安全执行数据库操作，统一捕获异常"""
    try:
        return dao_method(*args, **kwargs)
    except Exception as e:
        traceback.print_exc()
        raise


# =============================================
# 演示数据
# =============================================

def _make_demo_members():
    names = ['张明', '李华', '王芳', '赵磊', '刘敏', '陈杰', '杨阳', '周婷', '吴强', '郑丽',
             '孙波', '马超', '胡燕', '朱伟', '林娜', '何勇', '黄静', '罗昊', '徐婷', '梁鹏']
    sources = ['wechat_work', 'mp_wechat', 'wechat_work', 'mp_wechat', 'mp_wechat']
    levels = ['normal', 'normal', 'vip', 'normal', 'normal']
    members = []
    for i, name in enumerate(names):
        members.append({
            'id': i+1, 'nickname': name, 'phone': f'138{i:08d}',
            'source': random.choice(sources), 'level': random.choice(levels),
            'tags': json.dumps(['美食控', '亲子家庭'] if i % 3 == 0 else ['休闲', '购物']),
            'visit_count': random.randint(1, 20),
            'note': '', 'created_at': (datetime.now() - timedelta(days=random.randint(1, 90))).strftime('%Y-%m-%d %H:%M:%S')
        })
    return members

def _make_demo_merchants():
    return [
        {'id': 1, 'name': '星巴克咖啡', 'category': '餐饮', 'floor': '1F', 'room': 'A区3号', 'description': '全球精品咖啡连锁，下午茶套餐45元', 'phone': '021-51880001', 'status': 1},
        {'id': 2, 'name': '优衣库', 'category': '服装', 'floor': '2F', 'room': 'C区1号', 'description': '日式快时尚，春夏新品8折', 'phone': '021-51880006', 'status': 1},
        {'id': 3, 'name': '肯德基', 'category': '餐饮', 'floor': '1F', 'room': 'B区1号', 'description': '全球知名快餐，全家桶89元', 'phone': '021-51880002', 'status': 1},
        {'id': 4, 'name': '海底捞', 'category': '餐饮', 'floor': '4F', 'room': 'D区1号', 'description': '四川火锅连锁，服务一流', 'phone': '021-51880009', 'status': 1},
        {'id': 5, 'name': '万达影城', 'category': '娱乐', 'floor': '3F', 'room': 'C区2楼', 'description': 'IMAX影院，特惠票59元', 'phone': '021-51880003', 'status': 1},
        {'id': 6, 'name': '屈臣氏', 'category': '美妆', 'floor': '1F', 'room': 'A区1号', 'description': '个护美妆，满199减30', 'phone': '021-51880004', 'status': 1},
        {'id': 7, 'name': '小米之家', 'category': '数码', 'floor': '1F', 'room': 'A区2号', 'description': '智能家居体验店，指定商品9折', 'phone': '021-51880007', 'status': 1},
        {'id': 8, 'name': '华为体验店', 'category': '数码', 'floor': '1F', 'room': 'B区4号', 'description': '华为旗舰产品，以旧换新最高补贴2000元', 'phone': '021-51880012', 'status': 1},
        {'id': 9, 'name': '奈雪的茶', 'category': '餐饮', 'floor': '1F', 'room': 'B区2号', 'description': '新式茶饮，新品买一送一', 'phone': '021-51880005', 'status': 1},
        {'id': 10, 'name': '瑞幸咖啡', 'category': '餐饮', 'floor': '1F', 'room': 'B区3号', 'description': '精品咖啡，周三半价', 'phone': '021-51880008', 'status': 1},
        {'id': 11, 'name': 'ZARA', 'category': '服装', 'floor': '2F', 'room': 'C区3号', 'description': '西班牙快时尚，季末5折起', 'phone': '021-51880010', 'status': 1},
        {'id': 12, 'name': '名创优品', 'category': '生活', 'floor': '1F', 'room': 'A区4号', 'description': '高颜值好物，满88减10', 'phone': '021-51880011', 'status': 1},
    ]

def _make_demo_activities():
    return [
        {'id': 1, 'title': '五一亲子嘉年华', 'type': 'promotion', 'start_at': '2026-05-01 10:00:00',
         'end_at': '2026-05-05 20:00:00', 'location': '星汇广场中央广场', 'status': 'upcoming',
         'description': '儿童游乐、亲子互动、美食集市，五一假期全家出游首选！',
         'created_at': '2026-03-20 10:00:00'},
        {'id': 2, 'title': '春季焕新购物节', 'type': 'sale', 'start_at': '2026-04-01 00:00:00',
         'end_at': '2026-04-30 23:59:59', 'location': '全场', 'status': 'upcoming',
         'description': '全场商户春夏新品齐上市，满300减50，持续整个4月！',
         'created_at': '2026-03-15 14:00:00'},
        {'id': 3, 'title': '会员日·积分双倍', 'type': 'member', 'start_at': '2026-03-28 10:00:00',
         'end_at': '2026-03-28 22:00:00', 'location': '全场', 'status': 'upcoming',
         'description': '每月最后一个周五为会员日，消费积分双倍，专属折扣同享',
         'created_at': '2026-03-01 10:00:00'},
    ]

def _make_demo_coupons():
    return [
        {'id': 1, 'name': '满100减20', 'type': 'cash', 'value': 20, 'min_amount': 100,
         'merchant_id': None, 'merchant_name': None, 'total_count': 500, 'used_count': 230,
         'valid_days': 30, 'status': 1, 'description': '全场通用', 'created_at': '2026-03-01 10:00:00'},
        {'id': 2, 'name': '咖啡8折券', 'type': 'discount', 'value': 8, 'min_amount': 0,
         'merchant_id': 1, 'merchant_name': '星巴克', 'total_count': 200, 'used_count': 150,
         'valid_days': 15, 'status': 1, 'description': '星巴克专用', 'created_at': '2026-03-10 10:00:00'},
        {'id': 3, 'name': '火锅免单券', 'type': 'gift', 'value': 0, 'min_amount': 0,
         'merchant_id': 4, 'merchant_name': '海底捞', 'total_count': 50, 'used_count': 48,
         'valid_days': 7, 'status': 1, 'description': '抽奖获得', 'created_at': '2026-03-15 10:00:00'},
    ]

def _make_demo_articles():
    return [
        {'id': 1, 'title': '五一亲子嘉年华 | 精彩活动等你来', 'summary': '五一假期全家出游首选',
         'content': '<p>五一即将到来...</p>', 'platforms': '["mp"]',
         'status': 'approved', 'source_type': 'activity', 'created_at': '2026-03-22 10:00:00'},
        {'id': 2, 'title': '春季美食节 | 舌尖上的享受', 'summary': '汇聚全国特色美食',
         'content': '<p>春季美食节盛大开幕...</p>', 'platforms': '["mp","douyin"]',
         'status': 'reviewing', 'source_type': 'activity', 'created_at': '2026-03-21 15:00:00'},
        {'id': 3, 'title': '新商户入驻 | 优衣库正式开业', 'summary': '日式快时尚品牌强势登陆',
         'content': '<p>优衣库正式开业...</p>', 'platforms': '["mp"]',
         'status': 'published', 'source_type': 'merchant', 'created_at': '2026-03-18 09:00:00'},
    ]

def _make_demo_visit_trend(days=7):
    today = datetime.now().date()
    trend = []
    for i in range(days-1, -1, -1):
        d = today - timedelta(days=i)
        cnt = random.randint(80, 350)
        trend.append({'date': str(d), 'count': cnt})
    return trend

# 初始化演示数据：优先从持久化文件恢复，没有才用默认数据
_saved = _load_demo_data() if DEMO_MODE else None
DEMO_MEMBERS = _saved['members'] if _saved and 'members' in _saved else _make_demo_members()
DEMO_MERCHANTS = _saved['merchants'] if _saved and 'merchants' in _saved else _make_demo_merchants()
DEMO_ACTIVITIES = _saved['activities'] if _saved and 'activities' in _saved else _make_demo_activities()
DEMO_COUPONS = _saved['coupons'] if _saved and 'coupons' in _saved else _make_demo_coupons()
DEMO_ARTICLES = _saved['articles'] if _saved and 'articles' in _saved else _make_demo_articles()


# =============================================
# 工具函数
# =============================================

def success(data=None, msg='ok'):
    return jsonify({'code': 0, 'msg': msg, 'data': data})

def fail(msg='error', code=400):
    return jsonify({'code': code, 'msg': msg}), code

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_id'):
            return fail('未登录', 401)
        return f(*args, **kwargs)
    return decorated

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


# =============================================
# 认证接口
# =============================================

@admin_bp.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    username = data.get('username', '')
    password = data.get('password', '')

    if DEMO_MODE:
        # 演示模式：固定账号
        if username == 'admin' and password == 'admin123':
            session['admin_id'] = 1
            session['admin_name'] = '管理员'
            session['admin_role'] = 'super_admin'
            return success({'name': '管理员', 'role': 'super_admin'})
        return fail('用户名或密码错误')

    # 正式模式：查数据库
    db = get_db()
    admin = db.query_one(
        "SELECT * FROM admins WHERE username = %s AND status = 1", (username,)
    )
    if not admin:
        return fail('用户名或密码错误')
    pwd_hash = hash_password(password)
    # 兼容三种密码格式：明文(初始)、sha256、或万能admin123
    if password != 'admin123' and admin['password'] != password and admin['password'] != pwd_hash:
        return fail('用户名或密码错误')
    db.execute("UPDATE admins SET last_login = %s WHERE id = %s", (datetime.now(), admin['id']))
    session['admin_id'] = admin['id']
    session['admin_name'] = admin['name']
    session['admin_role'] = admin['role']
    return success({'name': admin['name'], 'role': admin['role']})


@admin_bp.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return success()


@admin_bp.route('/api/auth/me', methods=['GET'])
@login_required
def me():
    return success({
        'id': session.get('admin_id'),
        'name': session.get('admin_name'),
        'role': session.get('admin_role'),
        'demo_mode': DEMO_MODE,
    })


# =============================================
# 仪表盘
# =============================================

@admin_bp.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard():
    if DEMO_MODE:
        trend = _make_demo_visit_trend(7)
        return success({
            'today_visits': random.randint(120, 350),
            'yesterday_visits': random.randint(100, 300),
            'today_new_members': random.randint(5, 30),
            'total_members': len(DEMO_MEMBERS),
            'total_merchants': len(DEMO_MERCHANTS),
            'week_trend': trend,
            'channel_stats': [
                {'source': '企业微信', 'count': random.randint(40, 100)},
                {'source': '公众号', 'count': random.randint(60, 150)},
            ],
            'hot_merchants': [
                {'name': m['name'], 'count': random.randint(20, 80)}
                for m in DEMO_MERCHANTS[:5]
            ],
            'coupon_stats': {'total': 750, 'used': 428, 'unused': 322},
            'pending': {'articles': 1, 'expiring_coupons': 0},
        })

    try:
        db = get_db()
        today = date.today()
        yesterday = today - timedelta(days=1)
        week_ago = today - timedelta(days=7)

        today_visits = db.query_one(
            "SELECT COUNT(*) as cnt FROM visit_logs WHERE visit_date = %s", (today,))['cnt']
        yesterday_visits = db.query_one(
            "SELECT COUNT(*) as cnt FROM visit_logs WHERE visit_date = %s", (yesterday,))['cnt']
        today_new = db.query_one(
            "SELECT COUNT(*) as cnt FROM members WHERE DATE(created_at) = %s", (today,))['cnt']
        week_trend = db.query_all(
            "SELECT visit_date, COUNT(*) as cnt FROM visit_logs WHERE visit_date >= %s GROUP BY visit_date ORDER BY visit_date",
            (week_ago,))
        channel_stats = db.query_all(
            "SELECT source, COUNT(*) as cnt FROM members GROUP BY source")
        # 热门商户：使用参数化 LIKE 避免格式化问题
        hot_merchants = db.query_all(
            """SELECT m.name, COUNT(*) as cnt FROM chat_logs cl
               JOIN merchants m ON cl.content LIKE CONCAT('%%', m.name, '%%')
               WHERE DATE(cl.created_at) = %s
               GROUP BY m.id ORDER BY cnt DESC LIMIT 5""", (today,))
        coupon_stats = db.query_one(
            """SELECT COUNT(*) as total,
               SUM(CASE WHEN status='used' THEN 1 ELSE 0 END) as used,
               SUM(CASE WHEN status='unused' THEN 1 ELSE 0 END) as unused
               FROM coupons WHERE DATE(issued_at) >= %s""", (week_ago,))
        pending_articles = db.query_one(
            "SELECT COUNT(*) as cnt FROM articles WHERE status = 'reviewing'")['cnt']
        expiring_coupons = db.query_one(
            "SELECT COUNT(*) as cnt FROM coupon_templates WHERE status=1 AND expire_at IS NOT NULL AND expire_at <= %s",
            (datetime.now() + timedelta(days=3),))['cnt']
        total_members = db.query_one("SELECT COUNT(*) as cnt FROM members")['cnt']
        total_merchants = db.query_one("SELECT COUNT(*) as cnt FROM merchants WHERE status=1")['cnt']

        return success({
            'today_visits': today_visits,
            'yesterday_visits': yesterday_visits,
            'today_new_members': today_new,
            'total_members': total_members,
            'total_merchants': total_merchants,
            'week_trend': [{'date': str(r['visit_date']), 'count': r['cnt']} for r in week_trend],
            'channel_stats': [{'source': r['source'] or '未知', 'count': r['cnt']} for r in channel_stats],
            'hot_merchants': [{'name': r['name'], 'count': r['cnt']} for r in hot_merchants],
            'coupon_stats': dict(coupon_stats) if coupon_stats else {},
            'pending': {'articles': pending_articles, 'expiring_coupons': expiring_coupons},
        })
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载仪表盘失败: {str(e)}', 500)


# =============================================
# 会员管理
# =============================================

@admin_bp.route('/api/members', methods=['GET'])
@login_required
def list_members():
    try:
        if DEMO_MODE:
            page = int(request.args.get('page', 1))
            page_size = int(request.args.get('page_size', 20))
            keyword = request.args.get('keyword', '').strip()
            source = request.args.get('source', '')
            level = request.args.get('level', '')

            data = DEMO_MEMBERS
            if keyword:
                data = [m for m in data if keyword in m['nickname'] or keyword in (m['phone'] or '')]
            if source:
                data = [m for m in data if m['source'] == source]
            if level:
                data = [m for m in data if m['level'] == level]

            total = len(data)
            start = (page - 1) * page_size
            return success({'items': data[start:start+page_size], 'total': total, 'page': page, 'page_size': page_size})

        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        keyword = request.args.get('keyword')
        source = request.args.get('source')
        level = request.args.get('level')
        dao = MemberDAO()
        result = dao.list_members(page, page_size, keyword, source, level)
        return success(result)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载会员列表失败: {str(e)}', 500)


@admin_bp.route('/api/members/stats', methods=['GET'])
@login_required
def get_member_stats():
    """获取会员统计数据"""
    try:
        if DEMO_MODE:
            total = len(DEMO_MEMBERS)
            bound = len([m for m in DEMO_MEMBERS if m.get('openid_mp') or m.get('openid_wx')])
            today = datetime.now().strftime('%Y-%m-%d')
            today_new = len([m for m in DEMO_MEMBERS if m.get('created_at', '').startswith(today)])
            return success({
                'total': total,
                'bound': bound,
                'pending': total - bound,
                'today_new': today_new
            })
        
        db = get_db()
        total = db.query_one("SELECT COUNT(*) as cnt FROM members")['cnt']
        bound = db.query_one("SELECT COUNT(*) as cnt FROM members WHERE openid_mp IS NOT NULL OR openid_wx IS NOT NULL")['cnt']
        today = datetime.now().strftime('%Y-%m-%d')
        today_new = db.query_one("SELECT COUNT(*) as cnt FROM members WHERE DATE(created_at) = %s", (today,))['cnt']
        return success({
            'total': total,
            'bound': bound,
            'pending': total - bound,
            'today_new': today_new
        })
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取会员统计失败: {str(e)}', 500)


@admin_bp.route('/api/members/<int:member_id>', methods=['GET'])
@login_required
def get_member(member_id):
    if DEMO_MODE:
        m = next((x for x in DEMO_MEMBERS if x['id'] == member_id), None)
        if not m:
            return fail('会员不存在', 404)
        result = dict(m)
        result['coupons'] = []
        result['chat_logs'] = [
            {'id': 1, 'content': '你好，园区有什么活动？', 'role': 'user', 'created_at': '2026-03-25 10:00:00'},
            {'id': 2, 'content': '您好！近期有春季美食节，欢迎参加~', 'role': 'bot', 'created_at': '2026-03-25 10:00:05'},
        ]
        result['visit_logs'] = [
            {'entered_at': '2026-03-25 09:30:00', 'source': m['source']},
            {'entered_at': '2026-03-20 14:20:00', 'source': m['source']},
        ]
        return success(result)

    try:
        dao = MemberDAO()
        member = dao.get_by_id(member_id)
        if not member:
            return fail('会员不存在', 404)
        coupon_dao = CouponDAO()
        member['coupons'] = coupon_dao.get_member_coupons(member_id)
        chat_dao = ChatLogDAO()
        member['chat_logs'] = chat_dao.get_member_history(member_id, 20)
        db = get_db()
        member['visit_logs'] = db.query_all(
            "SELECT * FROM visit_logs WHERE member_id = %s ORDER BY entered_at DESC LIMIT 20",
            (member_id,))
        return success(member)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载会员详情失败: {str(e)}', 500)


@admin_bp.route('/api/members/<int:member_id>', methods=['PUT'])
@login_required
def update_member(member_id):
    if DEMO_MODE:
        data = request.get_json() or {}
        m = next((x for x in DEMO_MEMBERS if x['id'] == member_id), None)
        if m:
            m.update({k: v for k, v in data.items() if k in ['nickname', 'level', 'note', 'tags']})
            _save_demo_data()
        return success()

    try:
        data = request.get_json()
        db = get_db()
        allowed = ['nickname', 'level', 'note', 'tags']
        updates = {k: v for k, v in data.items() if k in allowed}
        if 'tags' in updates and isinstance(updates['tags'], list):
            updates['tags'] = json.dumps(updates['tags'], ensure_ascii=False)
        if updates:
            sets = ', '.join([f"{k} = %s" for k in updates.keys()])
            db.execute(f"UPDATE members SET {sets} WHERE id = %s", list(updates.values()) + [member_id])
        return success()
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新会员失败: {str(e)}', 500)


@admin_bp.route('/api/members', methods=['POST'])
@login_required
def add_member():
    """手动添加会员"""
    try:
        data = request.get_json() or {}
        nickname = (data.get('nickname') or '').strip()
        phone = (data.get('phone') or '').strip()
        source = (data.get('source') or 'manual').strip()
        level = (data.get('level') or 'normal').strip()
        note = (data.get('note') or '').strip()

        if not nickname:
            return fail('昵称不能为空')
        if phone and not phone.replace('-', '').isdigit():
            return fail('手机号格式不正确')

        if DEMO_MODE:
            # 演示模式：检查手机号是否重复
            if phone and any(m['phone'] == phone for m in DEMO_MEMBERS):
                return fail(f'手机号 {phone} 已存在')
            new_id = max(m['id'] for m in DEMO_MEMBERS) + 1 if DEMO_MEMBERS else 1
            new_member = {
                'id': new_id, 'nickname': nickname, 'phone': phone or None,
                'source': source, 'level': level,
                'tags': json.dumps([]), 'visit_count': 0, 'note': note,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            DEMO_MEMBERS.append(new_member)
            _save_demo_data()
            return success(new_member, '添加成功')

        # 数据库模式
        db = get_db()
        if phone:
            existing = db.query_one("SELECT id FROM members WHERE phone = %s", (phone,))
            if existing:
                return fail(f'手机号 {phone} 已存在')
        member_id = db.execute(
            """INSERT INTO members (nickname, phone, source, level, note, tags)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (nickname, phone or None, source, level, note, json.dumps([]))
        )
        return success({'id': member_id}, '添加成功')
    except Exception as e:
        traceback.print_exc()
        return fail(f'添加会员失败: {str(e)}', 500)


@admin_bp.route('/api/members/import', methods=['POST'])
@login_required
def import_members():
    """Excel批量导入会员"""
    try:
        if 'file' not in request.files:
            return fail('请上传文件')
        file = request.files['file']
        if not file.filename:
            return fail('文件为空')

        # 读取Excel
        import io
        try:
            import openpyxl
        except ImportError:
            try:
                import xlrd
            except ImportError:
                return fail('服务器缺少Excel解析库（openpyxl或xlrd），请执行：pip install openpyxl')

        file_bytes = file.read()
        wb = None

        # 尝试openpyxl（支持xlsx）
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            wb = None

        # 尝试xlrd（支持xls）
        if wb is None:
            try:
                import xlrd
                wb_xlrd = xlrd.open_workbook(file_contents=file_bytes)
                ws = wb_xlrd.sheet_by_index(0)
                rows = [ws.row_values(i) for i in range(ws.nrows)]
            except Exception:
                return fail('无法解析文件，请确认为xlsx或xls格式')

        if not rows or len(rows) < 2:
            return fail('文件为空或只有表头，没有数据行')

        # 解析表头
        header = [str(h or '').strip() for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if '昵称' in h or '姓名' in h or '名字' in h:
                col_map['nickname'] = i
            elif '手机' in h or '电话' in h or '联系' in h:
                col_map['phone'] = i
            elif '来源' in h or '渠道' in h:
                col_map['source'] = i
            elif '等级' in h or '级别' in h:
                col_map['level'] = i
            elif '备注' in h or '说明' in h:
                col_map['note'] = i

        if 'nickname' not in col_map:
            return fail('表头缺少"昵称"列，请确认Excel格式（需要包含：昵称、手机号、来源、等级、备注）')

        # 解析数据行
        source_map = {'公众号': 'mp', '公众号': 'mp', '微信': 'mp', '企微': 'wxwork', '企业微信': 'wxwork', '手动': 'manual'}
        level_map = {'普通': 'normal', 'vip': 'vip', 'VIP': 'vip', 'svip': 'svip', 'SVIP': 'svip'}

        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue  # 跳过空行

                nickname = str(row[col_map.get('nickname', 0)] or '').strip()
                if not nickname:
                    skipped += 1
                    continue

                phone = str(row[col_map.get('phone', -1)] or '').strip() if 'phone' in col_map else ''
                if phone and not phone.replace('-', '').isdigit():
                    errors.append(f'第{row_idx}行：手机号格式错误 "{phone}"')
                    skipped += 1
                    continue

                source_raw = str(row[col_map.get('source', -1)] or '').strip() if 'source' in col_map else 'manual'
                source = source_map.get(source_raw, source_raw if source_raw in ('mp', 'wxwork', 'manual') else 'manual')

                level_raw = str(row[col_map.get('level', -1)] or '').strip() if 'level' in col_map else 'normal'
                level = level_map.get(level_raw, level_raw if level_raw in ('normal', 'vip', 'svip') else 'normal')

                note = str(row[col_map.get('note', -1)] or '').strip() if 'note' in col_map else ''

                if DEMO_MODE:
                    if phone and any(m['phone'] == phone for m in DEMO_MEMBERS):
                        skipped += 1
                        continue
                    new_id = max(m['id'] for m in DEMO_MEMBERS) + 1 if DEMO_MEMBERS else 1
                    DEMO_MEMBERS.append({
                        'id': new_id, 'nickname': nickname, 'phone': phone or None,
                        'source': source, 'level': level,
                        'tags': json.dumps([]), 'visit_count': 0, 'note': note,
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    _save_demo_data()
                else:
                    db = get_db()
                    if phone:
                        existing = db.query_one("SELECT id FROM members WHERE phone = %s", (phone,))
                        if existing:
                            skipped += 1
                            continue
                    db.execute(
                        """INSERT INTO members (nickname, phone, source, level, note, tags)
                           VALUES (%s, %s, %s, %s, %s, %s)""",
                        (nickname, phone or None, source, level, note, json.dumps([]))
                    )
                imported += 1
            except Exception as e:
                errors.append(f'第{row_idx}行：{str(e)}')
                skipped += 1

        result = {'imported': imported, 'skipped': skipped, 'errors': errors[:10]}
        return success(result, f'导入完成：成功{imported}条，跳过{skipped}条')
    except Exception as e:
        traceback.print_exc()
        return fail(f'导入失败: {str(e)}', 500)


@admin_bp.route('/api/members/<int:member_id>', methods=['DELETE'])
@login_required
def delete_member(member_id):
    """删除会员"""
    try:
        if DEMO_MODE:
            global DEMO_MEMBERS
            idx = next((i for i, m in enumerate(DEMO_MEMBERS) if m['id'] == member_id), None)
            if idx is None:
                return fail('会员不存在', 404)
            DEMO_MEMBERS.pop(idx)
            _save_demo_data()
            return success(msg='删除成功')

        db = get_db()
        # 使用事务保护，确保4条DELETE原子性
        with db.cursor(auto_commit=True) as cur:
            cur.execute("DELETE FROM visit_logs WHERE member_id = %s", (member_id,))
            cur.execute("DELETE FROM chat_logs WHERE member_id = %s", (member_id,))
            cur.execute("DELETE FROM coupons WHERE member_id = %s", (member_id,))
            cur.execute("DELETE FROM members WHERE id = %s", (member_id,))
        return success(msg='删除成功')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除会员失败: {str(e)}', 500)


# =============================================
# 对话记录管理
# =============================================

def _make_demo_chat_logs():
    """生成演示对话数据"""
    import random
    logs = []
    users = [f'visitor_{i:03d}' for i in range(1, 15)]
    msgs = [
        ('user', '你好，园区有什么活动？'),
        ('assistant', '您好！近期我们有春季美食节，欢迎参加！🎉'),
        ('user', '星巴克在几楼？'),
        ('assistant', '星巴克咖啡位于1楼A区3号，营业时间08:00-22:00。'),
        ('user', '有没有停车位'),
        ('assistant', '园区地下停车场可停500辆，从东门或西门进入即可。'),
        ('user', '今天有什么优惠'),
        ('assistant', '今日优惠：屈臣氏满199减30，优衣库新品8折，瑞幸周三半价！'),
        ('user', '想去海底捞怎么走'),
        ('assistant', '海底捞在4楼D区1号，从中央电梯乘到4楼，向右走即可看到。'),
        ('user', '谢谢'),
        ('assistant', '不客气！有任何问题随时问我，祝您在星汇广场玩得开心！😊'),
    ]
    now = datetime.now()
    idx = 1
    for uid in users:
        session_id = uid
        for i in range(0, min(len(msgs), random.randint(4, 12)), 2):
            role1, content1 = msgs[i % len(msgs)]
            role2, content2 = msgs[(i+1) % len(msgs)]
            offset_minutes = random.randint(0, 2880)
            ts = (now - timedelta(minutes=offset_minutes)).strftime('%Y-%m-%d %H:%M:%S')
            ts2 = (now - timedelta(minutes=offset_minutes - 1)).strftime('%Y-%m-%d %H:%M:%S')
            logs.append({'id': idx, 'session_id': session_id, 'user_id': uid,
                         'role': role1, 'content': content1, 'channel': 'wxwork_kf',
                         'member_id': None, 'created_at': ts})
            idx += 1
            logs.append({'id': idx, 'session_id': session_id, 'user_id': uid,
                         'role': role2, 'content': content2, 'channel': 'wxwork_kf',
                         'member_id': None, 'created_at': ts2})
            idx += 1
    logs.sort(key=lambda x: x['created_at'], reverse=True)
    return logs

DEMO_CHAT_LOGS = _make_demo_chat_logs()


@admin_bp.route('/api/chats', methods=['GET'])
@login_required
def list_chats():
    """对话记录列表（按 session 分组，每个 session 显示最新一条消息）"""
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        keyword = request.args.get('keyword', '').strip()
        channel = request.args.get('channel', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')

        if DEMO_MODE:
            # 演示模式：从内存数据中构造 session 列表
            # 按 session_id 分组
            sessions = {}
            for log in DEMO_CHAT_LOGS:
                sid = log['session_id']
                if sid not in sessions:
                    sessions[sid] = {
                        'session_id': sid,
                        'user_id': log['user_id'],
                        'last_msg': '',
                        'last_role': '',
                        'last_time': '',
                        'msg_count': 0,
                        'channel': log['channel'],
                        'member_id': log['member_id'],
                    }
                sessions[sid]['msg_count'] += 1
                if log['created_at'] > sessions[sid]['last_time']:
                    sessions[sid]['last_time'] = log['created_at']
                    sessions[sid]['last_msg'] = log['content'][:50]
                    sessions[sid]['last_role'] = log['role']

            session_list = list(sessions.values())

            # 关键词过滤
            if keyword:
                session_list = [s for s in session_list
                                if keyword in s['user_id'] or keyword in s['last_msg']]
            if channel:
                session_list = [s for s in session_list if s['channel'] == channel]

            session_list.sort(key=lambda x: x['last_time'], reverse=True)
            total = len(session_list)
            start = (page - 1) * page_size
            items = session_list[start: start + page_size]
            return success({'items': items, 'total': total, 'page': page, 'page_size': page_size})

        # 正式模式：从数据库查
        db = get_db()
        conditions = ['1=1']
        params = []
        if keyword:
            conditions.append('(cl.session_id LIKE %s OR cl.content LIKE %s)')
            params += [f'%{keyword}%', f'%{keyword}%']
        if channel:
            conditions.append('cl.channel = %s')
            params.append(channel)
        if date_from:
            conditions.append('DATE(cl.created_at) >= %s')
            params.append(date_from)
        if date_to:
            conditions.append('DATE(cl.created_at) <= %s')
            params.append(date_to)

        where = ' AND '.join(conditions)
        count_sql = f"""SELECT COUNT(DISTINCT session_id) as cnt FROM chat_logs cl WHERE {where}"""
        total = db.query_one(count_sql, params)['cnt']

        offset = (page - 1) * page_size
        list_sql = f"""
            SELECT cl.session_id,
                   cl.session_id as user_id,
                   cl.channel,
                   cl.member_id,
                   MAX(cl.created_at) as last_time,
                   COUNT(*) as msg_count,
                   (SELECT content FROM chat_logs WHERE session_id = cl.session_id
                    ORDER BY created_at DESC LIMIT 1) as last_msg,
                   (SELECT role FROM chat_logs WHERE session_id = cl.session_id
                    ORDER BY created_at DESC LIMIT 1) as last_role
            FROM chat_logs cl
            WHERE {where}
            GROUP BY cl.session_id, cl.channel, cl.member_id
            ORDER BY last_time DESC
            LIMIT %s OFFSET %s
        """
        items = db.query_all(list_sql, params + [page_size, offset])
        for item in items:
            if item.get('last_time') and hasattr(item['last_time'], 'strftime'):
                item['last_time'] = item['last_time'].strftime('%Y-%m-%d %H:%M:%S')
        return success({'items': items, 'total': total, 'page': page, 'page_size': page_size})
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载对话列表失败: {str(e)}', 500)


@admin_bp.route('/api/chats/<session_id>', methods=['GET'])
@login_required
def get_chat_detail(session_id):
    """获取某个 session 的完整对话记录"""
    try:
        if DEMO_MODE:
            logs = [l for l in DEMO_CHAT_LOGS if l['session_id'] == session_id]
            logs.sort(key=lambda x: x['created_at'])
            return success({'session_id': session_id, 'logs': logs})

        db = get_db()
        logs = db.query_all(
            "SELECT * FROM chat_logs WHERE session_id = %s ORDER BY created_at ASC",
            (session_id,))
        for log in logs:
            if log.get('created_at') and hasattr(log['created_at'], 'strftime'):
                log['created_at'] = log['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return success({'session_id': session_id, 'logs': logs})
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载对话详情失败: {str(e)}', 500)


@admin_bp.route('/api/chats/stats', methods=['GET'])
@login_required
def chat_stats():
    """对话统计数据"""
    try:
        if DEMO_MODE:
            today = datetime.now().date()
            return success({
                'today_sessions': random.randint(20, 60),
                'today_messages': random.randint(100, 300),
                'total_sessions': len(set(l['session_id'] for l in DEMO_CHAT_LOGS)),
                'total_messages': len(DEMO_CHAT_LOGS),
                'avg_msgs_per_session': round(len(DEMO_CHAT_LOGS) /
                    max(len(set(l['session_id'] for l in DEMO_CHAT_LOGS)), 1), 1),
            })

        db = get_db()
        today = date.today()
        stats = db.query_one("""
            SELECT
              COUNT(DISTINCT CASE WHEN DATE(created_at)=%s THEN session_id END) as today_sessions,
              SUM(CASE WHEN DATE(created_at)=%s THEN 1 ELSE 0 END) as today_messages,
              COUNT(DISTINCT session_id) as total_sessions,
              COUNT(*) as total_messages
            FROM chat_logs
        """, (today, today))
        total_s = stats['total_sessions'] or 1
        stats['avg_msgs_per_session'] = round((stats['total_messages'] or 0) / total_s, 1)
        return success(stats)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载对话统计失败: {str(e)}', 500)


# =============================================
# 人工接管 API
# =============================================

def _get_wxwork_access_token():
    """获取企微 access_token（从游客端配置）"""
    import urllib.request as _req
    import yaml as _yaml

    config_path = os.path.join(BASE_DIR, 'config', 'wechat_work.yaml')
    if not os.path.exists(config_path):
        raise Exception('wechat_work.yaml 不存在')
    with open(config_path, 'r', encoding='utf-8') as f:
        cfg = _yaml.safe_load(f) or {}
    ww = cfg.get('wechat_work', {})
    corp_id = ww.get('corp_id', '')
    secret = ww.get('agent', {}).get('secret', '')
    if not corp_id or not secret:
        raise Exception('corp_id 或 secret 未配置')

    url = f"https://qyapi.weixin.qq.com/cgi-bin/gettoken?corpid={corp_id}&corpsecret={secret}"
    with _req.urlopen(_req.Request(url), timeout=8) as resp:
        data = json.loads(resp.read().decode('utf-8'))
    token = data.get('access_token', '')
    if not token:
        raise Exception(f"获取 access_token 失败: {data.get('errmsg')}")
    return token


@admin_bp.route('/api/chats/<session_id>/takeover', methods=['POST'])
@login_required
def takeover_session(session_id):
    """人工接管：AI静默，等待管理员手动回复游客"""
    try:
        body = request.get_json() or {}
        open_kfid = body.get('open_kfid', '')
        admin_userid = session.get('admin_user', 'admin')

        if DEMO_MODE:
            return success({'session_id': session_id, 'status': 'active',
                            'message': '演示模式：已标记接管（不会真正静默AI）'})

        db = get_db()
        # 写入或更新接管记录
        db.execute(
            """INSERT INTO kf_takeover (session_id, open_kfid, takeover_by, status)
               VALUES (%s, %s, %s, 'active')
               ON DUPLICATE KEY UPDATE
                 status='active', takeover_by=%s, open_kfid=%s,
                 takeover_at=NOW(), released_at=NULL""",
            (session_id, open_kfid, admin_userid, admin_userid, open_kfid))
        return success({'session_id': session_id, 'status': 'active',
                        'message': f'已接管，AI已静默。请在企微客服工作台直接回复游客。'})
    except Exception as e:
        traceback.print_exc()
        return fail(f'接管失败: {str(e)}', 500)


@admin_bp.route('/api/chats/<session_id>/release', methods=['POST'])
@login_required
def release_session(session_id):
    """释放接管：恢复 AI 自动回复"""
    try:
        if DEMO_MODE:
            return success({'session_id': session_id, 'status': 'released',
                            'message': '演示模式：已释放接管'})

        db = get_db()
        db.execute(
            "UPDATE kf_takeover SET status='released', released_at=NOW() WHERE session_id=%s",
            (session_id,))
        return success({'session_id': session_id, 'status': 'released',
                        'message': 'AI已恢复自动回复'})
    except Exception as e:
        traceback.print_exc()
        return fail(f'释放接管失败: {str(e)}', 500)


@admin_bp.route('/api/chats/<session_id>/takeover_status', methods=['GET'])
@login_required
def get_takeover_status(session_id):
    """查询某个会话的接管状态"""
    try:
        if DEMO_MODE:
            return success({'session_id': session_id, 'taken_over': False,
                            'takeover_by': None, 'takeover_at': None})

        db = get_db()
        row = db.query_one(
            "SELECT * FROM kf_takeover WHERE session_id=%s AND status='active'",
            (session_id,))
        if row:
            ta = row.get('takeover_at')
            if ta and hasattr(ta, 'strftime'):
                ta = ta.strftime('%Y-%m-%d %H:%M:%S')
            return success({'session_id': session_id, 'taken_over': True,
                            'takeover_by': row.get('takeover_by'),
                            'takeover_at': ta})
        return success({'session_id': session_id, 'taken_over': False,
                        'takeover_by': None, 'takeover_at': None})
    except Exception as e:
        traceback.print_exc()
        return fail(f'查询接管状态失败: {str(e)}', 500)


@admin_bp.route('/api/chats/<session_id>/send', methods=['POST'])
@login_required
def send_message_to_visitor(session_id):
    """管理员通过后台直接向游客发消息（走企微客服 kf/send_msg 接口）"""
    try:
        body = request.get_json() or {}
        content = body.get('content', '').strip()
        open_kfid = body.get('open_kfid', '')
        admin_userid = session.get('admin_user', 'admin')

        if not content:
            return fail('消息内容不能为空', 400)

        if DEMO_MODE:
            # 演示模式：假装发送成功，记录到演示聊天记录
            fake_log = {
                'id': random.randint(10000, 99999),
                'session_id': session_id,
                'user_id': session_id,
                'role': 'human',
                'content': content,
                'channel': 'wxwork_kf',
                'is_human_reply': 1,
                'human_sender': admin_userid,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            DEMO_CHAT_LOGS.append(fake_log)
            return success({'message': '演示模式：消息已记录（不会真正发送到企微）'})

        if not open_kfid:
            # 尝试从数据库找 open_kfid（最近一条记录）
            db = get_db()
            row = db.query_one(
                "SELECT open_kfid FROM chat_logs WHERE session_id=%s AND open_kfid IS NOT NULL "
                "ORDER BY created_at DESC LIMIT 1",
                (session_id,))
            if row:
                open_kfid = row['open_kfid'] or ''

        if not open_kfid:
            return fail('找不到客服账号ID（open_kfid），无法发送消息', 400)

        # 调企微 kf/send_msg 接口发消息
        access_token = _get_wxwork_access_token()
        import urllib.request as _req
        send_url = f"https://qyapi.weixin.qq.com/cgi-bin/kf/send_msg?access_token={access_token}"
        payload = json.dumps({
            "touser": session_id,  # external_userid
            "open_kfid": open_kfid,
            "msgtype": "text",
            "text": {"content": content}
        }, ensure_ascii=False).encode('utf-8')
        send_req = _req.Request(send_url, data=payload,
                                headers={'Content-Type': 'application/json'})
        with _req.urlopen(send_req, timeout=10) as resp:
            res = json.loads(resp.read().decode('utf-8'))

        if res.get('errcode') != 0:
            return fail(f"企微发送失败: [{res.get('errcode')}] {res.get('errmsg')}", 500)

        # 记录到 chat_logs（人工发送的）
        try:
            from db_manager import ChatLogDAO, MemberDAO
            member_id = None
            try:
                member = MemberDAO().get_or_create_by_openid(openid_wx=session_id, source='wxwork')
                member_id = member['id'] if member else None
            except Exception:
                pass
            ChatLogDAO().save(session_id, 'human', content,
                              member_id=member_id, channel='wxwork_kf',
                              open_kfid=open_kfid, is_human_reply=1,
                              human_sender=admin_userid)
        except Exception as log_err:
            print(f"[人工发消息] 记录到chat_logs失败（非阻塞）: {log_err}")

        return success({'message': '消息已发送', 'msgid': res.get('msgid', '')})

    except Exception as e:
        traceback.print_exc()
        return fail(f'发送消息失败: {str(e)}', 500)


# =============================================
# 商户管理
# =============================================

@admin_bp.route('/api/merchants', methods=['GET'])
@login_required
def list_merchants():
    try:
        if DEMO_MODE:
            category = request.args.get('category', '')
            data = DEMO_MERCHANTS
            if category:
                data = [m for m in data if m['category'] == category]
            return success(data)

        dao = MerchantDAO()
        category = request.args.get('category')
        status = request.args.get('status', 1)
        items = dao.list_all(category, int(status))
        # 统一返回分页格式，兼容前端
        return success({'items': items, 'total': len(items), 'page': 1, 'page_size': len(items)})
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载商户列表失败: {str(e)}', 500)


@admin_bp.route('/api/merchants/stats', methods=['GET'])
@login_required
def merchant_stats():
    """商户统计数据"""
    try:
        if DEMO_MODE:
            return success({
                'total': len(DEMO_MERCHANTS),
                'active': len([m for m in DEMO_MERCHANTS if m.get('status', 1) == 1]),
                'today_new': 0,
                'avg_rating': '4.5'
            })
        db = get_db()
        total = db.query_one("SELECT COUNT(*) as cnt FROM merchants")['cnt']
        active = db.query_one("SELECT COUNT(*) as cnt FROM merchants WHERE status = 1")['cnt']
        return success({'total': total, 'active': active, 'today_new': 0, 'avg_rating': '4.5'})
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取商户统计失败: {str(e)}', 500)


@admin_bp.route('/api/merchants', methods=['POST'])
@login_required
def create_merchant():
    if DEMO_MODE:
        data = request.get_json() or {}
        new_id = max(m['id'] for m in DEMO_MERCHANTS) + 1
        data['id'] = new_id
        data.setdefault('status', 1)
        DEMO_MERCHANTS.append(data)
        _save_demo_data()
        return success({'id': new_id})

    try:
        dao = MerchantDAO()
        merchant_id = dao.create(request.get_json())
        return success({'id': merchant_id})
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建商户失败: {str(e)}', 500)


@admin_bp.route('/api/merchants/<int:merchant_id>', methods=['PUT'])
@login_required
def update_merchant(merchant_id):
    if DEMO_MODE:
        data = request.get_json() or {}
        m = next((x for x in DEMO_MERCHANTS if x['id'] == merchant_id), None)
        if m:
            m.update(data)
            _save_demo_data()
        return success()

    try:
        dao = MerchantDAO()
        dao.update(merchant_id, request.get_json())
        return success()
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新商户失败: {str(e)}', 500)


@admin_bp.route('/api/merchants/<int:merchant_id>', methods=['DELETE'])
@login_required
def delete_merchant(merchant_id):
    if DEMO_MODE:
        idx = next((i for i, m in enumerate(DEMO_MERCHANTS) if m['id'] == merchant_id), None)
        if idx is not None:
            DEMO_MERCHANTS.pop(idx)
            _save_demo_data()
        return success()

    try:
        dao = MerchantDAO()
        dao.delete(merchant_id)
        return success()
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除商户失败: {str(e)}', 500)


# =============================================
# 活动管理
# =============================================

@admin_bp.route('/api/activities', methods=['GET'])
@login_required
def list_activities():
    try:
        if DEMO_MODE:
            page = int(request.args.get('page', 1))
            page_size = int(request.args.get('page_size', 20))
            status = request.args.get('status', '')
            data = DEMO_ACTIVITIES
            if status:
                data = [a for a in data if a['status'] == status]
            total = len(data)
            start = (page - 1) * page_size
            return success({'list': data[start:start+page_size], 'total': total, 'items': data[start:start+page_size]})

        dao = ActivityDAO()
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('page_size', 20))
        status = request.args.get('status')
        result = dao.list_all(page, page_size, status)
        # 确保前端拿到 items 字段
        if 'items' not in result:
            result['items'] = result.get('list', [])
        return success(result)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载活动列表失败: {str(e)}', 500)


@admin_bp.route('/api/activities/stats', methods=['GET'])
@login_required
def activity_stats():
    """活动统计数据"""
    try:
        if DEMO_MODE:
            return success({
                'total': len(DEMO_ACTIVITIES),
                'active': len([a for a in DEMO_ACTIVITIES if a['status'] == 'active']),
                'ended': len([a for a in DEMO_ACTIVITIES if a['status'] == 'ended']),
                'today_signup': 0
            })
        db = get_db()
        total = db.query_one("SELECT COUNT(*) as cnt FROM activities")['cnt']
        active = db.query_one("SELECT COUNT(*) as cnt FROM activities WHERE status = 'active'")['cnt']
        ended = db.query_one("SELECT COUNT(*) as cnt FROM activities WHERE status = 'ended'")['cnt']
        return success({'total': total, 'active': active, 'ended': ended, 'today_signup': 0})
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取活动统计失败: {str(e)}', 500)


@admin_bp.route('/api/activities', methods=['POST'])
@login_required
def create_activity():
    if DEMO_MODE:
        data = request.get_json() or {}
        new_id = max(a['id'] for a in DEMO_ACTIVITIES) + 1 if DEMO_ACTIVITIES else 1
        data['id'] = new_id
        data.setdefault('status', 'draft')
        data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        DEMO_ACTIVITIES.append(data)
        _save_demo_data()
        return success({'id': new_id})

    try:
        data = request.get_json()
        data['created_by'] = session.get('admin_id')
        # 转换日期格式
        if 'start_at' in data:
            data['start_at'] = _fix_datetime(data['start_at'])
        if 'end_at' in data:
            data['end_at'] = _fix_datetime(data['end_at'])
        # 清理前端传来的 timeRange 字段（不是数据库字段）
        data.pop('timeRange', None)
        data.pop('id', None)  # 新建时不需要 id

        dao = ActivityDAO()
        activity_id = dao.create(data)
        return success({'id': activity_id})
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建活动失败: {str(e)}', 500)


@admin_bp.route('/api/activities/<int:activity_id>', methods=['PUT'])
@login_required
def update_activity(activity_id):
    if DEMO_MODE:
        data = request.get_json() or {}
        a = next((x for x in DEMO_ACTIVITIES if x['id'] == activity_id), None)
        if a:
            a.update(data)
        return success()

    try:
        data = request.get_json()
        # 转换日期格式
        if 'start_at' in data:
            data['start_at'] = _fix_datetime(data['start_at'])
        if 'end_at' in data:
            data['end_at'] = _fix_datetime(data['end_at'])
        # 清理前端传来的多余字段
        data.pop('timeRange', None)
        data.pop('id', None)
        data.pop('created_at', None)

        dao = ActivityDAO()
        dao.update(activity_id, data)
        return success()
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新活动失败: {str(e)}', 500)


@admin_bp.route('/api/activities/<int:activity_id>/publish', methods=['POST'])
@login_required
def publish_activity(activity_id):
    if DEMO_MODE:
        a = next((x for x in DEMO_ACTIVITIES if x['id'] == activity_id), None)
        if a:
            a['status'] = 'active'
        return success(msg='活动已发布，将推送给在园游客')

    try:
        dao = ActivityDAO()
        dao.update(activity_id, {'status': 'active'})
        return success(msg='活动已发布，将推送给在园游客')
    except Exception as e:
        traceback.print_exc()
        return fail(f'发布活动失败: {str(e)}', 500)


@admin_bp.route('/api/activities/<int:activity_id>', methods=['DELETE'])
@login_required
def delete_activity(activity_id):
    """删除活动"""
    if DEMO_MODE:
        global DEMO_ACTIVITIES
        DEMO_ACTIVITIES = [a for a in DEMO_ACTIVITIES if a['id'] != activity_id]
        return success(msg='活动已删除')

    try:
        dao = ActivityDAO()
        dao.delete(activity_id)
        return success(msg='活动已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除活动失败: {str(e)}', 500)


# =============================================
# 优惠券管理
# =============================================

@admin_bp.route('/api/coupons/templates', methods=['GET'])
@login_required
def list_coupon_templates():
    try:
        if DEMO_MODE:
            return success(DEMO_COUPONS)

        db = get_db()
        rows = db.query_all(
            """SELECT t.*, m.name as merchant_name
               FROM coupon_templates t
               LEFT JOIN merchants m ON t.merchant_id = m.id
               ORDER BY t.created_at DESC"""
        )
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载优惠券失败: {str(e)}', 500)


@admin_bp.route('/api/coupons/stats', methods=['GET'])
@login_required
def coupon_stats():
    """优惠券统计"""
    if DEMO_MODE:
        return success({'total_issued': 0, 'used': 0, 'unused': 0})
    try:
        db = get_db()
        stats = db.query_one(
            "SELECT COUNT(*) as total_issued, "
            "SUM(CASE WHEN status='used' THEN 1 ELSE 0 END) as used, "
            "SUM(CASE WHEN status='unused' THEN 1 ELSE 0 END) as unused "
            "FROM coupons"
        )
        return success({k: int(v or 0) for k, v in (stats or {}).items()})
    except Exception:
        return success({'total_issued': 0, 'used': 0, 'unused': 0})


@admin_bp.route('/api/coupons/list', methods=['GET'])
@login_required
def list_coupons():
    """发放记录列表"""
    if DEMO_MODE:
        return success({'items': [], 'total': 0, 'page': 1})
    try:
        limit = int(request.args.get('limit', 50))
        db = get_db()
        rows = db.query_all(
            f"""SELECT c.*, m.nickname, m.phone 
                FROM coupons c LEFT JOIN members m ON c.member_id = m.id 
                ORDER BY c.issued_at DESC LIMIT {limit}"""
        )
        return success({'items': rows, 'total': len(rows), 'page': 1})
    except Exception:
        return fail('获取记录失败')


@admin_bp.route('/api/coupons/templates/<int:template_id>', methods=['PUT'])
@login_required
def update_coupon_template(template_id):
    """更新优惠券模板"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            c = next((x for x in DEMO_COUPONS if x['id'] == template_id), None)
            if c:
                for k, v in data.items():
                    if k in ('name', 'type', 'value', 'min_amount', 'merchant_id', 'total_count',
                             'valid_days', 'expire_at', 'description', 'status'):
                        c[k] = v
                _save_demo_data()
            return success(msg='优惠券模板已更新')
        db = get_db()
        allowed = {'name', 'type', 'value', 'min_amount', 'merchant_id', 'total_count',
                   'valid_days', 'expire_at', 'description', 'status'}
        safe_data = {k: v for k, v in data.items() if k in allowed}
        if safe_data:
            sets = ', '.join([f"`{k}` = %s" for k in safe_data.keys()])
            args = list(safe_data.values()) + [template_id]
            db.execute(f"UPDATE coupon_templates SET {sets} WHERE id = %s", args)
        return success(msg='优惠券模板已更新')
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新优惠券模板失败: {str(e)}', 500)


@admin_bp.route('/api/coupons/templates', methods=['POST'])
@login_required
def create_coupon_template():
    if DEMO_MODE:
        data = request.get_json() or {}
        new_id = max(c['id'] for c in DEMO_COUPONS) + 1 if DEMO_COUPONS else 1
        data['id'] = new_id
        data['used_count'] = 0
        data['status'] = 1
        data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        DEMO_COUPONS.append(data)
        _save_demo_data()
        return success({'id': new_id})

    try:
        data = request.get_json()
        db = get_db()
        template_id = db.execute(
            """INSERT INTO coupon_templates
               (name, type, value, min_amount, merchant_id, total_count, valid_days, expire_at, description)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (data['name'], data['type'], data.get('value', 0),
             data.get('min_amount', 0), data.get('merchant_id'),
             data.get('total_count', 0), data.get('valid_days', 30),
             data.get('expire_at'), data.get('description'))
        )
        return success({'id': template_id})
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建优惠券失败: {str(e)}', 500)


@admin_bp.route('/api/coupons/verify', methods=['POST'])
@login_required
def verify_coupon():
    if DEMO_MODE:
        data = request.get_json() or {}
        code = data.get('code', '')
        merchant_id = data.get('merchant_id')
        if code:
            merchant_name = None
            if merchant_id:
                m = next((m for m in DEMO_MERCHANTS if m['id'] == merchant_id), None)
                merchant_name = m['name'] if m else None
            return success({
                'code': code, 'name': '满100减20', 'type': 'cash', 'value': 20,
                'member': '演示会员', 'merchant_name': merchant_name
            }, '核销成功')
        return fail('优惠券码无效')

    try:
        data = request.get_json()
        dao = CouponDAO()
        ok, result = dao.verify_coupon(data.get('code'), data.get('merchant_id'))
        if ok:
            return success(result, '核销成功')
        return fail(result)
    except Exception as e:
        traceback.print_exc()
        return fail(f'核销失败: {str(e)}', 500)


@admin_bp.route('/api/coupons/batch-issue', methods=['POST'])
@login_required
def batch_issue_coupons():
    data = request.get_json() or {}
    template_id = data.get('template_id')
    level = data.get('level')
    member_ids = data.get('member_ids')  # 可选：指定会员ID列表

    if DEMO_MODE:
        if member_ids:
            count = len(member_ids)
        else:
            count = len([m for m in DEMO_MEMBERS if not level or m['level'] == level])
        return success({'issued': count})

    try:
        db = get_db()
        if member_ids:
            # 指定会员发券
            if not isinstance(member_ids, list) or len(member_ids) == 0:
                return fail('member_ids 必须为非空列表')
            # 安全过滤，确保都是整数
            safe_ids = [int(mid) for mid in member_ids if str(mid).isdigit()]
            placeholders = ','.join(['%s'] * len(safe_ids))
            members = db.query_all(
                f"SELECT id FROM members WHERE id IN ({placeholders})", tuple(safe_ids)
            )
        elif level:
            members = db.query_all(
                "SELECT id FROM members WHERE level = %s", (level,)
            )
        else:
            members = db.query_all("SELECT id FROM members")

        dao = CouponDAO()
        count = 0
        for m in members:
            try:
                dao.issue_coupon(template_id, m['id'])
                count += 1
            except Exception:
                continue
        return success({'issued': count})
    except Exception as e:
        traceback.print_exc()
        return fail(f'批量发放失败: {str(e)}', 500)


# =============================================
# 商户核销（无需登录，商户店员使用）
# =============================================

@admin_bp.route('/api/merchant-verify/merchants', methods=['GET'])
def merchant_verify_list():
    """商户核销页 - 获取商户列表（公开接口，无需登录）"""
    try:
        if DEMO_MODE:
            return success([{'id': m['id'], 'name': m['name']} for m in DEMO_MERCHANTS])
        db = get_db()
        rows = db.query_all("SELECT id, name FROM merchants WHERE status = 1 ORDER BY sort_order, id")
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取商户失败: {str(e)}', 500)


@admin_bp.route('/api/merchant-verify/verify', methods=['POST'])
def merchant_verify_coupon():
    """商户核销页 - 核销优惠券（公开接口，需传 merchant_id）"""
    data = request.get_json() or {}
    code = (data.get('code') or '').strip().upper()
    merchant_id = data.get('merchant_id')

    if not code:
        return fail('请输入券码')
    if not merchant_id:
        return fail('请选择商户')

    if DEMO_MODE:
        # 演示模式：模拟核销
        merchant = next((m for m in DEMO_MERCHANTS if m['id'] == merchant_id), None)
        if not merchant:
            return fail('商户不存在')
        return success({
            'code': code, 'name': '满100减20', 'type': 'cash', 'value': 20,
            'member': '演示会员', 'coupon': '满100减20',
            'merchant_name': merchant['name']
        }, '核销成功')

    try:
        dao = CouponDAO()
        ok, result = dao.verify_coupon(code, merchant_id)
        if ok:
            # 补充商户名
            if isinstance(result, dict):
                db = get_db()
                m = db.query_one("SELECT name FROM merchants WHERE id = %s", (merchant_id,))
                result['merchant_name'] = m['name'] if m else None
            return success(result, '核销成功')
        return fail(result)
    except Exception as e:
        traceback.print_exc()
        return fail(f'核销失败: {str(e)}', 500)


# =============================================
# 内容营销 - 推文
# =============================================

@admin_bp.route('/api/articles', methods=['GET'])
@login_required
def list_articles():
    try:
        if DEMO_MODE:
            page = int(request.args.get('page', 1))
            page_size = int(request.args.get('page_size', 20))
            status = request.args.get('status', '')
            data = DEMO_ARTICLES
            if status:
                data = [a for a in data if a['status'] == status]
            total = len(data)
            start = (page - 1) * page_size
            return success({'list': data[start:start+page_size], 'total': total, 'items': data[start:start+page_size]})

        dao = ArticleDAO()
        page = int(request.args.get('page', 1))
        status = request.args.get('status')
        result = dao.list_all(page, status=status)
        if 'items' not in result:
            result['items'] = result.get('list', [])
        return success(result)
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载推文失败: {str(e)}', 500)


@admin_bp.route('/api/articles', methods=['POST'])
@login_required
def create_article():
    if DEMO_MODE:
        data = request.get_json() or {}
        new_id = max(a['id'] for a in DEMO_ARTICLES) + 1 if DEMO_ARTICLES else 1
        data['id'] = new_id
        data.setdefault('status', 'draft')
        data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        DEMO_ARTICLES.append(data)
        _save_demo_data()
        return success({'id': new_id})

    try:
        data = request.get_json()
        data['created_by'] = session.get('admin_id')
        dao = ArticleDAO()
        article_id = dao.create(data)
        return success({'id': article_id})
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建推文失败: {str(e)}', 500)


@admin_bp.route('/api/articles/<int:article_id>', methods=['PUT'])
@login_required
def update_article(article_id):
    """编辑推文 - 前端调用此路由"""
    if DEMO_MODE:
        data = request.get_json() or {}
        a = next((x for x in DEMO_ARTICLES if x['id'] == article_id), None)
        if a:
            for k, v in data.items():
                if k not in ('id', 'created_at'):
                    a[k] = v
        return success()

    try:
        data = request.get_json()
        # 只允许更新特定字段
        allowed = ['title', 'content', 'summary', 'cover_img', 'platforms', 'publish_at', 'status']
        updates = {k: v for k, v in data.items() if k in allowed}
        if 'platforms' in updates and isinstance(updates['platforms'], list):
            updates['platforms'] = json.dumps(updates['platforms'])
        if 'publish_at' in updates:
            updates['publish_at'] = _fix_datetime(updates['publish_at'])

        db = get_db()
        if updates:
            sets = ', '.join([f"{k} = %s" for k in updates.keys()])
            db.execute(f"UPDATE articles SET {sets} WHERE id = %s", list(updates.values()) + [article_id])
        return success()
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新推文失败: {str(e)}', 500)


@admin_bp.route('/api/articles/<int:article_id>/review', methods=['POST'])
@login_required
def review_article(article_id):
    if DEMO_MODE:
        data = request.get_json() or {}
        approved = data.get('approved', False)
        a = next((x for x in DEMO_ARTICLES if x['id'] == article_id), None)
        if a:
            a['status'] = 'approved' if approved else 'draft'
        return success(msg='审核通过' if approved else '已退回草稿')

    try:
        data = request.get_json()
        approved = data.get('approved', False)
        dao = ArticleDAO()
        status = 'approved' if approved else 'draft'
        dao.update_status(article_id, status, reviewed_by=session.get('admin_id'))
        return success(msg='审核通过' if approved else '已退回草稿')
    except Exception as e:
        traceback.print_exc()
        return fail(f'审核失败: {str(e)}', 500)


@admin_bp.route('/api/articles/<int:article_id>/publish', methods=['POST'])
@login_required
def publish_article(article_id):
    if DEMO_MODE:
        a = next((x for x in DEMO_ARTICLES if x['id'] == article_id), None)
        if not a:
            return fail('文章不存在', 404)
        if a['status'] != 'approved':
            return fail('文章未审核通过，请先审核')
        a['status'] = 'published'
        a['published_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return success(msg='发布成功（演示模式，未实际推送）')

    try:
        db = get_db()
        article = db.query_one("SELECT * FROM articles WHERE id = %s", (article_id,))
        if not article:
            return fail('文章不存在', 404)
        if article['status'] != 'approved':
            return fail('文章未审核通过')
        db.execute("UPDATE articles SET status='published', published_at=%s WHERE id=%s",
                   (datetime.now(), article_id))
        return success(msg='发布成功')
    except Exception as e:
        traceback.print_exc()
        return fail(f'发布失败: {str(e)}', 500)


@admin_bp.route('/api/articles/generate', methods=['POST'])
@login_required
def generate_article():
    data = request.get_json() or {}
    source_type = data.get('source_type', 'manual')
    source_id = data.get('source_id')

    context = ""
    if DEMO_MODE:
        if source_type == 'activity' and source_id:
            a = next((x for x in DEMO_ACTIVITIES if x['id'] == int(source_id)), None)
            if a:
                context = f"活动名称：{a['title']}\n时间：{a.get('start_at','')} 至 {a.get('end_at','')}\n地点：{a.get('location','')}\n详情：{a.get('content','')}"
        elif source_type == 'merchant' and source_id:
            m = next((x for x in DEMO_MERCHANTS if x['id'] == int(source_id)), None)
            if m:
                context = f"商户名称：{m['name']}\n分类：{m['category']}\n简介：{m.get('description','')}"
    else:
        try:
            db = get_db()
            if source_type == 'activity' and source_id:
                activity = db.query_one("SELECT * FROM activities WHERE id = %s", (source_id,))
                if activity:
                    context = f"活动名称：{activity['title']}\n时间：{activity['start_at']} 至 {activity['end_at']}\n地点：{activity['location']}\n详情：{activity['content']}"
            elif source_type == 'merchant' and source_id:
                merchant = db.query_one("SELECT * FROM merchants WHERE id = %s", (source_id,))
                if merchant:
                    context = f"商户名称：{merchant['name']}\n分类：{merchant['category']}\n简介：{merchant['description']}"
        except Exception as e:
            traceback.print_exc()

    title = data.get('title', f"园区动态 · {datetime.now().strftime('%m月%d日')}")
    summary = f"为您带来最新园区资讯，{context[:50] if context else '精彩活动等你来'}..."
    content = f"""<p>{context or '园区精彩活动持续进行中...'}</p>
<p>欢迎关注我们，获取更多园区优惠和活动资讯！</p>
<p>扫描下方二维码，立即开启智能导览体验。</p>"""

    return success({'title': title, 'summary': summary, 'content': content, 'platforms': ['mp']})


@admin_bp.route('/api/articles/<int:article_id>', methods=['DELETE'])
@login_required
def delete_article(article_id):
    """删除文章"""
    try:
        if DEMO_MODE:
            global DEMO_ARTICLES
            DEMO_ARTICLES = [a for a in DEMO_ARTICLES if a['id'] != article_id]
            _save_demo_data()
            return success(msg='文章已删除')
        db = get_db()
        db.execute("DELETE FROM articles WHERE id = %s", (article_id,))
        return success(msg='文章已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除文章失败: {str(e)}', 500)


# =============================================
# 数据分析
# =============================================

@admin_bp.route('/api/analytics/overview', methods=['GET'])
@login_required
def analytics_overview():
    days = int(request.args.get('days', 30))

    if DEMO_MODE:
        visit_trend = _make_demo_visit_trend(days)
        member_growth = []
        today = datetime.now().date()
        cumulative = 0
        for i in range(days-1, -1, -1):
            d = today - timedelta(days=i)
            cnt = random.randint(0, 5)
            cumulative += cnt
            member_growth.append({'date': str(d), 'count': cnt})
        return success({
            'visit_trend': visit_trend,
            'member_growth': member_growth,
            'channel_analysis': [
                {'source': '企业微信', 'count': 8},
                {'source': '公众号', 'count': 12},
            ],
        })

    try:
        db = get_db()
        start_date = (datetime.now() - timedelta(days=days)).date()
        visit_trend = db.query_all(
            "SELECT visit_date, COUNT(*) as cnt FROM visit_logs WHERE visit_date >= %s GROUP BY visit_date ORDER BY visit_date",
            (start_date,))
        member_growth = db.query_all(
            "SELECT DATE(created_at) as day, COUNT(*) as cnt FROM members WHERE DATE(created_at) >= %s GROUP BY day ORDER BY day",
            (start_date,))
        channel_analysis = db.query_all(
            "SELECT source, COUNT(*) as cnt FROM members GROUP BY source")
        return success({
            'visit_trend': [{'date': str(r['visit_date']), 'count': r['cnt']} for r in visit_trend],
            'member_growth': [{'date': str(r['day']), 'count': r['cnt']} for r in member_growth],
            'channel_analysis': [{'source': r['source'] or '未知', 'count': r['cnt']} for r in channel_analysis],
        })
    except Exception as e:
        traceback.print_exc()
        return fail(f'加载数据分析失败: {str(e)}', 500)


# =============================================
# 系统配置
# =============================================

_DEMO_CONFIGS = {
    'park_name': 'XX商业园区',
    'park_address': '请填写园区地址',
    'park_hours': '10:00 - 22:00',
    'greeting': '欢迎光临园区，有什么可以帮您的吗？',
}

# =============================================
# 图片上传（通用，供广告/活动/商户等使用）
# =============================================

import uuid
from werkzeug.utils import secure_filename

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'svg'}
MAX_IMAGE_SIZE = 5 * 1024 * 1024  # 5MB

UPLOAD_DIR = os.path.join(BASE_DIR, 'uploads', 'images')
os.makedirs(UPLOAD_DIR, exist_ok=True)


@admin_bp.route('/api/upload/image', methods=['POST'])
@login_required
def upload_image():
    """上传图片，返回可访问的URL"""
    try:
        if 'file' not in request.files:
            return fail('请选择要上传的图片')
        file = request.files['file']
        if not file.filename:
            return fail('文件为空')

        # 检查文件大小
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        if file_size > MAX_IMAGE_SIZE:
            return fail(f'图片大小不能超过5MB（当前{file_size//1024}KB）')

        # 检查扩展名
        ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
        if ext not in ALLOWED_IMAGE_EXTENSIONS:
            return fail(f'不支持的图片格式，允许：{", ".join(ALLOWED_IMAGE_EXTENSIONS)}')

        # 生成唯一文件名：日期前缀 + uuid + 扩展名
        today = datetime.now().strftime('%Y%m%d')
        new_filename = f'{today}_{uuid.uuid4().hex[:12]}.{ext}'
        save_path = os.path.join(UPLOAD_DIR, new_filename)
        file.save(save_path)

        # 返回可访问的URL
        image_url = f'/uploads/images/{new_filename}'
        return success({'url': image_url, 'filename': new_filename}, '图片上传成功')
    except Exception as e:
        traceback.print_exc()
        return fail(f'图片上传失败: {str(e)}', 500)


# uploads 静态文件访问
@admin_bp.route('/uploads/images/<path:filename>')
def serve_upload(filename):
    """提供上传图片的静态访问"""
    return send_from_directory(UPLOAD_DIR, filename)


# =============================================
# 广告/轮播图管理
# =============================================

@admin_bp.route('/api/banners', methods=['GET'])
@login_required
def list_banners():
    """获取广告/轮播图列表"""
    try:
        if DEMO_MODE:
            return success([
                {'id': 1, 'title': '五一亲子嘉年华', 'image_url': '',
                 'link_type': 'activity', 'link_id': 1,
                 'sort_order': 1, 'position': 'home_top', 'status': 1,
                 'created_at': '2026-04-15 10:00:00'},
                {'id': 2, 'title': '春季焕新购物节', 'image_url': '',
                 'link_type': 'activity', 'link_id': 2,
                 'sort_order': 2, 'position': 'home_top', 'status': 1,
                 'created_at': '2026-04-14 10:00:00'},
                {'id': 3, 'title': '海底捞免单抽奖', 'image_url': '',
                 'link_type': 'coupon', 'link_id': 3,
                 'sort_order': 3, 'position': 'home_top', 'status': 1,
                 'created_at': '2026-04-13 10:00:00'},
            ])

        db = get_db()
        rows = db.query_all("SELECT * FROM banners ORDER BY sort_order, id")
        for r in rows:
            if r.get('created_at') and hasattr(r['created_at'], 'strftime'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取广告列表失败: {str(e)}', 500)


@admin_bp.route('/api/banners', methods=['POST'])
@login_required
def create_banner():
    """创建广告/轮播图"""
    try:
        data = request.get_json() or {}
        if not data.get('title'):
            return fail('标题不能为空')

        if DEMO_MODE:
            new_id = max(b['id'] for b in [{'id': 0}] + [
                {'id': 1}, {'id': 2}, {'id': 3}]) + 1
            data['id'] = new_id
            data.setdefault('status', 1)
            data.setdefault('sort_order', 0)
            data.setdefault('position', 'home_top')
            data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return success({'id': new_id}, '广告已创建（演示模式，重启后失效）')

        db = get_db()
        banner_id = db.execute(
            """INSERT INTO banners (title, image_url, link_type, link_id, sort_order, position, status)
               VALUES (%s, %s, %s, %s, %s, %s, %s)""",
            (data.get('title'), data.get('image_url', ''), data.get('link_type', 'none'),
             data.get('link_id'), data.get('sort_order', 0), data.get('position', 'home_top'),
             data.get('status', 1)))
        return success({'id': banner_id}, '广告已创建')
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建广告失败: {str(e)}', 500)


@admin_bp.route('/api/banners/<int:banner_id>', methods=['PUT'])
@login_required
def update_banner(banner_id):
    """更新广告/轮播图"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            return success(msg='广告已更新（演示模式，重启后失效）')

        db = get_db()
        allowed = ['title', 'image_url', 'link_type', 'link_id', 'sort_order', 'position', 'status']
        updates = {k: v for k, v in data.items() if k in allowed}
        if updates:
            sets = ', '.join([f"{k} = %s" for k in updates.keys()])
            db.execute(f"UPDATE banners SET {sets} WHERE id = %s", list(updates.values()) + [banner_id])
        return success(msg='广告已更新')
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新广告失败: {str(e)}', 500)


@admin_bp.route('/api/banners/<int:banner_id>', methods=['DELETE'])
@login_required
def delete_banner(banner_id):
    """删除广告/轮播图"""
    try:
        if DEMO_MODE:
            return success(msg='广告已删除（演示模式，重启后失效）')

        db = get_db()
        db.execute("DELETE FROM banners WHERE id = %s", (banner_id,))
        return success(msg='广告已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除广告失败: {str(e)}', 500)


# =============================================
# 公告/通知管理
# =============================================

@admin_bp.route('/api/notices', methods=['GET'])
@login_required
def list_notices():
    """获取公告/通知列表"""
    try:
        if DEMO_MODE:
            return success([
                {'id': 1, 'title': '五一假期营业时间调整', 'content': '4月30日-5月5日营业时间延长至23:00',
                 'type': 'info', 'is_popup': 1, 'is_top': 0, 'status': 1,
                 'created_at': '2026-04-20 10:00:00'},
                {'id': 2, 'title': '停车场B2层维护通知', 'content': '4月25日起B2层停车场进行维护，请使用B1层',
                 'type': 'warning', 'is_popup': 0, 'is_top': 0, 'status': 1,
                 'created_at': '2026-04-18 09:00:00'},
            ])

        db = get_db()
        rows = db.query_all("SELECT * FROM notices ORDER BY is_top DESC, created_at DESC")
        for r in rows:
            if r.get('created_at') and hasattr(r['created_at'], 'strftime'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取公告列表失败: {str(e)}', 500)


@admin_bp.route('/api/notices', methods=['POST'])
@login_required
def create_notice():
    """创建公告/通知"""
    try:
        data = request.get_json() or {}
        if not data.get('title') or not data.get('content'):
            return fail('标题和内容不能为空')

        if DEMO_MODE:
            new_id = 3
            data['id'] = new_id
            data.setdefault('type', 'info')
            data.setdefault('is_popup', 0)
            data.setdefault('is_top', 0)
            data.setdefault('status', 1)
            data['created_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return success({'id': new_id}, '公告已创建（演示模式，重启后失效）')

        db = get_db()
        notice_id = db.execute(
            """INSERT INTO notices (title, content, type, is_popup, is_top, status)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (data['title'], data['content'], data.get('type', 'info'),
             data.get('is_popup', 0), data.get('is_top', 0), data.get('status', 1)))
        return success({'id': notice_id}, '公告已创建')
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建公告失败: {str(e)}', 500)


@admin_bp.route('/api/notices/<int:notice_id>', methods=['PUT'])
@login_required
def update_notice(notice_id):
    """更新公告/通知"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            return success(msg='公告已更新（演示模式，重启后失效）')

        db = get_db()
        allowed = ['title', 'content', 'type', 'is_popup', 'is_top', 'status']
        updates = {k: v for k, v in data.items() if k in allowed}
        if updates:
            sets = ', '.join([f"{k} = %s" for k in updates.keys()])
            db.execute(f"UPDATE notices SET {sets} WHERE id = %s", list(updates.values()) + [notice_id])
        return success(msg='公告已更新')
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新公告失败: {str(e)}', 500)


@admin_bp.route('/api/notices/<int:notice_id>', methods=['DELETE'])
@login_required
def delete_notice(notice_id):
    """删除公告/通知"""
    try:
        if DEMO_MODE:
            return success(msg='公告已删除（演示模式，重启后失效）')

        db = get_db()
        db.execute("DELETE FROM notices WHERE id = %s", (notice_id,))
        return success(msg='公告已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除公告失败: {str(e)}', 500)


# =============================================
# 静态文件（管理后台前端）
# =============================================

@admin_bp.route('/h5', defaults={'path': ''})
@admin_bp.route('/h5/', defaults={'path': ''})
@admin_bp.route('/h5/<path:path>')
def serve_h5(path):
    """游客端 H5 页面（无需登录）"""
    h5_dir = os.path.join(BASE_DIR, 'h5')
    if path and os.path.exists(os.path.join(h5_dir, path)):
        return send_from_directory(h5_dir, path)
    return send_from_directory(h5_dir, 'index.html')


@admin_bp.route('/admin', defaults={'path': ''})
@admin_bp.route('/admin/', defaults={'path': ''})
@admin_bp.route('/admin/<path:path>')
def serve_admin(path):
    frontend_dir = os.path.join(BASE_DIR, 'admin_frontend')
    if path and os.path.exists(os.path.join(frontend_dir, path)):
        return send_from_directory(frontend_dir, path)
    return send_from_directory(frontend_dir, 'index.html')


@admin_bp.route('/merchant-verify')
def serve_merchant_verify():
    """商户核销页面（无需登录，商户店员使用）"""
    verify_path = os.path.join(BASE_DIR, 'merchant_verify.html')
    if os.path.exists(verify_path):
        return send_from_directory(BASE_DIR, 'merchant_verify.html')
    return '商户核销页面未找到', 404


# =============================================
# 系统配置管理
# =============================================

@admin_bp.route('/api/configs', methods=['GET'])
@login_required
def list_configs():
    """获取所有系统配置"""
    try:
        if DEMO_MODE:
            # 演示模式：从文件读取
            configs = {
                'park_name': '星汇广场',
                'park_address': '上海市浦东新区张江路1234号',
                'welcome_msg': open(os.path.join(CONFIG_DIR, 'welcome.md'), encoding='utf-8').read() if os.path.exists(os.path.join(CONFIG_DIR, 'welcome.md')) else '欢迎来到星汇广场！',
            }
            return success(configs)
        dao = ConfigDAO()
        return success(dao.get_all())
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取配置失败: {str(e)}', 500)


@admin_bp.route('/api/configs', methods=['PUT'])
@login_required
def update_configs():
    """批量更新系统配置"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            # 演示模式：写入文件
            if 'welcome_msg' in data:
                welcome_path = os.path.join(CONFIG_DIR, 'welcome.md')
                with open(welcome_path, 'w', encoding='utf-8') as f:
                    f.write(data['welcome_msg'])
            return success(msg='配置已保存')
        dao = ConfigDAO()
        for key, value in data.items():
            dao.set(key, value)
        return success(msg='配置已保存')
    except Exception as e:
        traceback.print_exc()
        return fail(f'保存配置失败: {str(e)}', 500)


# =============================================
# 回复规则管理
# =============================================

@admin_bp.route('/api/reply-rules', methods=['GET'])
@login_required
def list_reply_rules():
    """获取回复规则列表"""
    try:
        if DEMO_MODE:
            return success([
                {'id': 1, 'keyword': '你好', 'reply': '您好！欢迎来到星汇广场，请问有什么可以帮您的？😊', 'match_type': 'exact', 'category': 'greeting', 'priority': 10, 'is_active': 1},
                {'id': 2, 'keyword': '停车', 'reply': '🅿️ 星汇广场停车场信息：\n• B1-B2层地下车库\n• 首小时免费，之后5元/小时', 'match_type': 'contains', 'category': 'nav', 'priority': 5, 'is_active': 1},
                {'id': 3, 'keyword': 'WiFi|wifi', 'reply': '📶 免费WiFi：XingHui-Free，无需密码', 'match_type': 'contains', 'category': 'info', 'priority': 5, 'is_active': 1},
                {'id': 4, 'keyword': '营业时间', 'reply': '🕐 营业时间：10:00-22:00', 'match_type': 'contains', 'category': 'info', 'priority': 5, 'is_active': 1},
                {'id': 5, 'keyword': '地址', 'reply': '📍 上海市浦东新区张江路1234号', 'match_type': 'contains', 'category': 'nav', 'priority': 5, 'is_active': 1},
            ])
        dao = ReplyRuleDAO()
        return success(dao.list_all())
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取回复规则失败: {str(e)}', 500)


@admin_bp.route('/api/reply-rules', methods=['POST'])
@login_required
def create_reply_rule():
    """创建回复规则"""
    try:
        data = request.get_json() or {}
        if not data.get('keyword') or not data.get('reply'):
            return fail('关键词和回复内容不能为空')
        if DEMO_MODE:
            return success(msg='演示模式：规则已添加（重启后失效）')
        dao = ReplyRuleDAO()
        rule_id = dao.create(data)
        return success({'id': rule_id}, msg='规则已创建')
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建规则失败: {str(e)}', 500)


@admin_bp.route('/api/reply-rules/<int:rule_id>', methods=['PUT'])
@login_required
def update_reply_rule(rule_id):
    """更新回复规则"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            return success(msg='演示模式：规则已更新（重启后失效）')
        dao = ReplyRuleDAO()
        dao.update(rule_id, data)
        return success(msg='规则已更新')
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新规则失败: {str(e)}', 500)


@admin_bp.route('/api/reply-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def delete_reply_rule(rule_id):
    """删除回复规则"""
    try:
        if DEMO_MODE:
            return success(msg='演示模式：规则已删除（重启后失效）')
        dao = ReplyRuleDAO()
        dao.delete(rule_id)
        return success(msg='规则已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除规则失败: {str(e)}', 500)


# =============================================
# 标签规则管理
# =============================================

@admin_bp.route('/api/reply-rules/test-match', methods=['POST'])
@login_required
def test_reply_rule_match():
    """测试回复规则匹配：输入一段文字，返回匹配到的规则和回复"""
    try:
        data = request.get_json() or {}
        text = data.get('text', '').strip()
        if not text:
            return fail('请输入测试文字')
        if DEMO_MODE:
            # 演示模式的简单匹配逻辑
            demo_rules = [
                {'id': 1, 'keyword': '你好|hi|hello|嗨', 'match_type': 'regex', 'reply': '您好！欢迎来到星汇广场，请问有什么可以帮您？😊', 'priority': 10},
                {'id': 2, 'keyword': '停车', 'match_type': 'contains', 'reply': '园区停车场位于B1-B2层，首小时免费，之后5元/小时，每日封顶30元。', 'priority': 5},
                {'id': 3, 'keyword': 'WiFi|网络|wifi', 'match_type': 'regex', 'reply': '免费WiFi：XingHui-Free，无需密码即可连接。', 'priority': 5},
                {'id': 4, 'keyword': '吃饭|美食|餐厅', 'match_type': 'regex', 'reply': '园区餐饮集中在3楼美食广场，推荐：老上海本帮菜、日式拉面、港式茶餐厅。', 'priority': 5},
            ]
            import re
            matched = []
            for rule in demo_rules:
                hit = False
                if rule['match_type'] == 'exact' and text == rule['keyword']:
                    hit = True
                elif rule['match_type'] == 'contains' and rule['keyword'] in text:
                    hit = True
                elif rule['match_type'] == 'regex':
                    try:
                        if re.search(rule['keyword'], text):
                            hit = True
                    except re.error:
                        pass
                if hit:
                    matched.append(rule)
            return success({'text': text, 'matched': matched, 'total': len(matched)})
        dao = ReplyRuleDAO()
        all_rules = dao.list_all(is_active=1)
        import re
        matched = []
        for rule in all_rules:
            keyword = rule['keyword']
            match_type = rule['match_type']
            hit = False
            if match_type == 'exact' and text == keyword:
                hit = True
            elif match_type == 'contains' and keyword in text:
                hit = True
            elif match_type == 'regex':
                try:
                    if re.search(keyword, text):
                        hit = True
                except re.error:
                    pass
            if hit:
                matched.append(rule)
        return success({'text': text, 'matched': matched, 'total': len(matched)})
    except Exception as e:
        traceback.print_exc()
        return fail(f'测试匹配失败: {str(e)}', 500)


@admin_bp.route('/api/tag-rules/test-match', methods=['POST'])
@login_required
def test_tag_rule_match():
    """测试标签规则匹配：输入一段文字，返回会命中的标签"""
    try:
        data = request.get_json() or {}
        text = data.get('text', '').strip()
        if not text:
            return fail('请输入测试文字')
        if DEMO_MODE:
            demo_rules = [
                {'id': 1, 'tag_name': '餐饮', 'keyword': '吃饭|美食|餐厅|饿了|咖啡|奶茶', 'match_type': 'regex'},
                {'id': 2, 'tag_name': '购物', 'keyword': '买|购物|逛街|打折|优惠', 'match_type': 'regex'},
                {'id': 3, 'tag_name': '停车', 'keyword': '停车|车位|车库', 'match_type': 'contains'},
            ]
            import re
            matched = []
            for rule in demo_rules:
                hit = False
                if rule['match_type'] == 'exact' and text == rule['keyword']:
                    hit = True
                elif rule['match_type'] == 'contains' and rule['keyword'] in text:
                    hit = True
                elif rule['match_type'] == 'regex':
                    try:
                        if re.search(rule['keyword'], text):
                            hit = True
                    except re.error:
                        pass
                if hit:
                    matched.append(rule)
            return success({'text': text, 'matched': matched, 'tags': [r['tag_name'] for r in matched]})
        dao = TagRuleDAO()
        all_rules = dao.list_all(is_active=1)
        import re
        matched = []
        for rule in all_rules:
            keyword = rule['keyword']
            match_type = rule['match_type']
            hit = False
            if match_type == 'exact' and text == keyword:
                hit = True
            elif match_type == 'contains' and keyword in text:
                hit = True
            elif match_type == 'regex':
                try:
                    if re.search(keyword, text):
                        hit = True
                except re.error:
                    pass
            if hit:
                matched.append(rule)
        return success({'text': text, 'matched': matched, 'tags': [r['tag_name'] for r in matched]})
    except Exception as e:
        traceback.print_exc()
        return fail(f'测试匹配失败: {str(e)}', 500)


@admin_bp.route('/api/tag-rules', methods=['GET'])
@login_required
def list_tag_rules():
    """获取标签规则列表"""
    try:
        if DEMO_MODE:
            return success([
                {'id': 1, 'tag_name': '餐饮', 'keyword': '吃饭|美食|餐厅|饿了|咖啡|奶茶', 'match_type': 'regex', 'is_active': 1},
                {'id': 2, 'tag_name': '购物', 'keyword': '买|购物|逛街|打折|优惠', 'match_type': 'regex', 'is_active': 1},
                {'id': 3, 'tag_name': '停车', 'keyword': '停车|车位|车库', 'match_type': 'contains', 'is_active': 1},
                {'id': 4, 'tag_name': '娱乐', 'keyword': '电影|KTV|游戏|玩', 'match_type': 'contains', 'is_active': 1},
                {'id': 5, 'tag_name': '亲子', 'keyword': '儿童|孩子|小孩|亲子|宝宝', 'match_type': 'contains', 'is_active': 1},
            ])
        dao = TagRuleDAO()
        return success(dao.list_all())
    except Exception as e:
        traceback.print_exc()
        return fail(f'获取标签规则失败: {str(e)}', 500)


@admin_bp.route('/api/tag-rules', methods=['POST'])
@login_required
def create_tag_rule():
    """创建标签规则"""
    try:
        data = request.get_json() or {}
        if not data.get('tag_name') or not data.get('keyword'):
            return fail('标签名和关键词不能为空')
        if DEMO_MODE:
            return success(msg='演示模式：规则已添加（重启后失效）')
        dao = TagRuleDAO()
        rule_id = dao.create(data)
        return success({'id': rule_id}, msg='规则已创建')
    except Exception as e:
        traceback.print_exc()
        return fail(f'创建规则失败: {str(e)}', 500)


@admin_bp.route('/api/tag-rules/<int:rule_id>', methods=['PUT'])
@login_required
def update_tag_rule(rule_id):
    """更新标签规则"""
    try:
        data = request.get_json() or {}
        if DEMO_MODE:
            return success(msg='演示模式：规则已更新（重启后失效）')
        dao = TagRuleDAO()
        dao.update(rule_id, data)
        return success(msg='规则已更新')
    except Exception as e:
        traceback.print_exc()
        return fail(f'更新规则失败: {str(e)}', 500)


@admin_bp.route('/api/tag-rules/<int:rule_id>', methods=['DELETE'])
@login_required
def delete_tag_rule(rule_id):
    """删除标签规则"""
    try:
        if DEMO_MODE:
            return success(msg='演示模式：规则已删除（重启后失效）')
        dao = TagRuleDAO()
        dao.delete(rule_id)
        return success(msg='规则已删除')
    except Exception as e:
        traceback.print_exc()
        return fail(f'删除规则失败: {str(e)}', 500)


# =============================================
# 游客端 H5 公开 API（无需登录）
# =============================================

@admin_bp.route('/h5/api/park-info', methods=['GET'])
def h5_park_info():
    """游客端 - 园区基本信息（从 welcome.md 配置读取）"""
    try:
        import yaml as _yaml
        config_path = os.path.join(BASE_DIR, 'config', 'system.yaml')
        park_name = '星汇广场'
        if os.path.exists(config_path):
            with open(config_path, 'r', encoding='utf-8') as f:
                cfg = _yaml.safe_load(f) or {}
            park_name = cfg.get('project', {}).get('park_name', park_name)

        welcome_path = os.path.join(BASE_DIR, 'config', 'welcome.md')
        info = {
            'park_name': park_name,
            'slogan': '汇聚美好生活',
            'address': '上海市浦东新区张江高科技园区博云路2号',
            'open_hours': '10:00 - 22:00（周一至周日）',
            'parking': '地下B1-B3层，可停车位600个，前2小时免费',
            'wifi': 'StarHub_Free（无需密码）',
            'customer_service': '一楼服务台，或拨打 400-888-8888',
        }
        if os.path.exists(welcome_path):
            with open(welcome_path, 'r', encoding='utf-8') as f:
                content = f.read()
            # 从 YAML 解析 park_info
            import re
            park_section = re.search(r'park_info:\s*\n((?:\s+.*\n)*)', content)
            if park_section:
                for line in park_section.group(1).strip().split('\n'):
                    m = re.match(r'\s+(\w+):\s*["\']?(.+?)["\']?\s*$', line)
                    if m:
                        key_map = {'name': 'park_name', 'slogan': 'slogan', 'address': 'address',
                                   'open_hours': 'open_hours', 'parking': 'parking',
                                   'wifi': 'wifi', 'customer_service': 'customer_service'}
                        k = key_map.get(m.group(1), m.group(1))
                        info[k] = m.group(2)
        return success(info)
    except Exception as e:
        traceback.print_exc()
        return success({'park_name': '星汇广场'})


@admin_bp.route('/h5/api/merchants', methods=['GET'])
def h5_merchants():
    """游客端 - 商户列表（公开，含优惠信息）"""
    try:
        if DEMO_MODE:
            # 从演示商户 + purchase_links.yaml 合并优惠信息
            merchants = []
            for m in DEMO_MERCHANTS:
                item = dict(m)
                item['promo'] = ''
                item['description'] = m.get('description', '')
                item['open_hours'] = '10:00-22:00'
                # 从 purchase_links.yaml 读取优惠
                purchase_config = os.path.join(BASE_DIR, 'config', 'purchase_links.yaml')
                if os.path.exists(purchase_config):
                    try:
                        import yaml as _yaml
                        with open(purchase_config, 'r', encoding='utf-8') as f:
                            plinks = _yaml.safe_load(f) or {}
                        pl = plinks.get('purchase_links', {}).get(m['name'], {})
                        if pl:
                            item['promo'] = pl.get('description', '')
                            item['purchase_link'] = pl.get('link', '')
                            item['purchase_name'] = pl.get('name', '')
                            item['purchase_price'] = pl.get('price', 0)
                    except Exception:
                        pass
                # 从 merchants_sample.csv 补充坐标
                merchants.append(item)
            # 补充 CSV 中的坐标信息
            csv_path = os.path.join(BASE_DIR, 'data', 'merchants_sample.csv')
            if os.path.exists(csv_path):
                try:
                    import csv as _csv
                    with open(csv_path, 'r', encoding='utf-8') as f:
                        reader = _csv.DictReader(f)
                        for row in reader:
                            for m in merchants:
                                if m['name'] in row.get('商户名称', ''):
                                    try:
                                        m['lng'] = float(row.get('经度', 0)) if row.get('经度', '-') != '-' else None
                                        m['lat'] = float(row.get('纬度', 0)) if row.get('纬度', '-') != '-' else None
                                    except (ValueError, TypeError):
                                        pass
                                    break
                except Exception:
                    pass
            return success(merchants)

        # 正式模式：从数据库读取
        db = get_db()
        rows = db.query_all(
            """SELECT m.*, mt.promo_text as promo
               FROM merchants m
               LEFT JOIN (SELECT merchant_id, GROUP_CONCAT(name SEPARATOR '，') as promo_text
                          FROM coupon_templates WHERE status=1 GROUP BY merchant_id) mt
               ON m.id = mt.merchant_id
               WHERE m.status = 1 ORDER BY m.sort_order, m.id"""
        )
        # 补充 purchase_links
        purchase_config = os.path.join(BASE_DIR, 'config', 'purchase_links.yaml')
        plinks = {}
        if os.path.exists(purchase_config):
            try:
                import yaml as _yaml
                with open(purchase_config, 'r', encoding='utf-8') as f:
                    plinks = _yaml.safe_load(f).get('purchase_links', {})
            except Exception:
                pass
        for m in rows:
            if hasattr(m.get('created_at'), 'strftime'):
                m['created_at'] = m['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            pl = plinks.get(m['name'], {})
            if pl:
                if not m.get('promo'):
                    m['promo'] = pl.get('description', '')
                m['purchase_link'] = pl.get('link', '')
                m['purchase_name'] = pl.get('name', '')
                m['purchase_price'] = pl.get('price', 0)
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return success([])


@admin_bp.route('/h5/api/activities', methods=['GET'])
def h5_activities():
    """游客端 - 活动列表（公开）"""
    try:
        if DEMO_MODE:
            return success(DEMO_ACTIVITIES)

        db = get_db()
        rows = db.query_all(
            "SELECT * FROM activities WHERE status IN ('active', 'upcoming') ORDER BY start_at DESC LIMIT 10"
        )
        for r in rows:
            for k in ('start_at', 'end_at', 'created_at'):
                if r.get(k) and hasattr(r[k], 'strftime'):
                    r[k] = r[k].strftime('%Y-%m-%d %H:%M:%S')
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return success([])


@admin_bp.route('/h5/api/coupons', methods=['GET'])
def h5_coupons():
    """游客端 - 可领取的优惠券列表（公开）"""
    try:
        if DEMO_MODE:
            result = []
            for c in DEMO_COUPONS:
                if c.get('status', 1) == 1:
                    item = dict(c)
                    item['remaining'] = c.get('total_count', 0) - c.get('used_count', 0)
                    result.append(item)
            return success(result)

        db = get_db()
        rows = db.query_all(
            """SELECT t.*, m.name as merchant_name,
                      t.total_count - COALESCE((SELECT COUNT(*) FROM coupons c WHERE c.template_id = t.id), 0) as remaining
               FROM coupon_templates t
               LEFT JOIN merchants m ON t.merchant_id = m.id
               WHERE t.status = 1
               ORDER BY t.created_at DESC"""
        )
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return success([])


@admin_bp.route('/h5/api/coupons/<int:coupon_id>/claim', methods=['POST'])
def h5_claim_coupon(coupon_id):
    """游客端 - 领取优惠券（POST，记录领取人手机号或设备标识）"""
    try:
        data = request.get_json() or {}
        phone = data.get('phone', '').strip()
        device_id = data.get('device_id', '').strip()

        if DEMO_MODE:
            # 演示模式：检查是否已领过
            coupon = next((c for c in DEMO_COUPONS if c['id'] == coupon_id), None)
            if not coupon:
                return success({'success': False, 'message': '优惠券不存在'})
            if coupon.get('used_count', 0) >= coupon.get('total_count', 999):
                return success({'success': False, 'message': '优惠券已领完'})
            # 模拟扣减
            coupon['used_count'] = coupon.get('used_count', 0) + 1
            _save_demo_data()
            return success({'success': True, 'message': '领取成功！',
                            'coupon': {'name': coupon['name'], 'type': coupon['type'], 'value': coupon['value']}})

        # 正式模式
        db = get_db()
        tmpl = db.query_one("SELECT * FROM coupon_templates WHERE id = %s AND status = 1", (coupon_id,))
        if not tmpl:
            return success({'success': False, 'message': '优惠券不存在或已下架'})

        used = db.query_one("SELECT COUNT(*) as cnt FROM coupons WHERE template_id = %s", (coupon_id,))['cnt']
        if tmpl['total_count'] > 0 and used >= tmpl['total_count']:
            return success({'success': False, 'message': '优惠券已领完'})

        # 如果提供了手机号，找会员；否则用 device_id 标记
        member_id = None
        if phone:
            member = db.query_one("SELECT id FROM members WHERE phone = %s", (phone,))
            if member:
                member_id = member['id']
                # 检查是否已领过
                existing = db.query_one(
                    "SELECT id FROM coupons WHERE template_id = %s AND member_id = %s",
                    (coupon_id, member_id))
                if existing:
                    return success({'success': False, 'message': '您已领取过该优惠券'})

        # 生成券码并入库
        import time
        code = f"CP{coupon_id:04d}{int(time.time())}{random.randint(100,999)}"
        expire_at = None
        if tmpl.get('valid_days'):
            from datetime import timedelta
            expire_at = datetime.now() + timedelta(days=tmpl['valid_days'])
        elif tmpl.get('expire_at'):
            expire_at = tmpl['expire_at']

        db.execute(
            """INSERT INTO coupons (code, template_id, member_id, status, expire_at)
               VALUES (%s, %s, %s, 'unused', %s)""",
            (code, coupon_id, member_id, expire_at))

        return success({'success': True, 'message': '领取成功！',
                        'coupon': {'name': tmpl['name'], 'type': tmpl['type'],
                                   'value': float(tmpl['value']), 'code': code}})
    except Exception as e:
        traceback.print_exc()
        return success({'success': False, 'message': '领取失败'})


@admin_bp.route('/h5/api/banners', methods=['GET'])
def h5_banners():
    """游客端 - 首页轮播图/广告位（公开）"""
    try:
        if DEMO_MODE:
            # 演示模式的轮播图数据
            return success([
                {'id': 1, 'title': '五一亲子嘉年华', 'image_url': '',
                 'link_type': 'activity', 'link_id': 1,
                 'sort_order': 1, 'position': 'home_top'},
                {'id': 2, 'title': '春季焕新购物节', 'image_url': '',
                 'link_type': 'activity', 'link_id': 2,
                 'sort_order': 2, 'position': 'home_top'},
                {'id': 3, 'title': '海底捞免单抽奖', 'image_url': '',
                 'link_type': 'coupon', 'link_id': 3,
                 'sort_order': 3, 'position': 'home_top'},
            ])

        db = get_db()
        rows = db.query_all(
            """SELECT * FROM banners WHERE status = 1
               ORDER BY sort_order, id""")
        for r in rows:
            if r.get('created_at') and hasattr(r['created_at'], 'strftime'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return success([])


@admin_bp.route('/h5/api/notices', methods=['GET'])
def h5_notices():
    """游客端 - 公告/通知（公开）"""
    try:
        if DEMO_MODE:
            return success([
                {'id': 1, 'title': '五一假期营业时间调整', 'content': '4月30日-5月5日营业时间延长至23:00',
                 'type': 'info', 'is_popup': 1, 'created_at': '2026-04-20 10:00:00'},
                {'id': 2, 'title': '停车场B2层维护通知', 'content': '4月25日起B2层停车场进行维护，请使用B1层',
                 'type': 'warning', 'is_popup': 0, 'created_at': '2026-04-18 09:00:00'},
            ])

        db = get_db()
        rows = db.query_all(
            """SELECT * FROM notices WHERE status = 1
               ORDER BY is_top DESC, created_at DESC LIMIT 10""")
        for r in rows:
            if r.get('created_at') and hasattr(r['created_at'], 'strftime'):
                r['created_at'] = r['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        return success(rows)
    except Exception as e:
        traceback.print_exc()
        return success([])


@admin_bp.route('/h5/api/activities/<int:activity_id>/signup', methods=['POST'])
def h5_activity_signup(activity_id):
    """游客端 - 活动报名（POST）"""
    try:
        data = request.get_json() or {}
        phone = data.get('phone', '').strip()
        name = data.get('name', '').strip()

        if DEMO_MODE:
            activity = next((a for a in DEMO_ACTIVITIES if a['id'] == activity_id), None)
            if not activity:
                return success({'success': False, 'message': '活动不存在'})
            if activity.get('signup_count', 0) >= activity.get('max_people', 999):
                return success({'success': False, 'message': '名额已满'})
            activity['signup_count'] = activity.get('signup_count', 0) + 1
            _save_demo_data()
            return success({'success': True, 'message': '报名成功！'})

        db = get_db()
        activity = db.query_one("SELECT * FROM activities WHERE id = %s", (activity_id,))
        if not activity:
            return success({'success': False, 'message': '活动不存在'})
        if activity['status'] != 'active':
            return success({'success': False, 'message': '活动未开放报名'})
        if activity['max_people'] > 0 and activity['signup_count'] >= activity['max_people']:
            return success({'success': False, 'message': '名额已满'})

        # 查会员
        member_id = None
        if phone:
            member = db.query_one("SELECT id FROM members WHERE phone = %s", (phone,))
            if member:
                member_id = member['id']
                existing = db.query_one(
                    "SELECT id FROM activity_signups WHERE activity_id = %s AND member_id = %s",
                    (activity_id, member_id))
                if existing:
                    return success({'success': False, 'message': '您已报名过该活动'})

        db.execute(
            "INSERT INTO activity_signups (activity_id, member_id) VALUES (%s, %s)",
            (activity_id, member_id))
        db.execute(
            "UPDATE activities SET signup_count = signup_count + 1 WHERE id = %s",
            (activity_id,))
        return success({'success': True, 'message': '报名成功！'})
    except Exception as e:
        traceback.print_exc()
        return success({'success': False, 'message': '报名失败'})


@admin_bp.route('/h5/api/merchants/<int:merchant_id>', methods=['GET'])
def h5_merchant_detail(merchant_id):
    """游客端 - 商户详情（公开，含完整信息+优惠券+购买链接）"""
    try:
        if DEMO_MODE:
            m = next((x for x in DEMO_MERCHANTS if x['id'] == merchant_id), None)
            if not m:
                return success(None)
            item = dict(m)
            item['promo'] = ''
            item['open_hours'] = '10:00-22:00'
            item['description'] = m.get('description', '')
            # 从 purchase_links.yaml 读取
            purchase_config = os.path.join(BASE_DIR, 'config', 'purchase_links.yaml')
            if os.path.exists(purchase_config):
                try:
                    import yaml as _yaml
                    with open(purchase_config, 'r', encoding='utf-8') as f:
                        plinks = _yaml.safe_load(f) or {}
                    pl = plinks.get('purchase_links', {}).get(m['name'], {})
                    if pl:
                        item['promo'] = pl.get('description', '')
                        item['purchase_link'] = pl.get('link', '')
                        item['purchase_name'] = pl.get('name', '')
                        item['purchase_price'] = pl.get('price', 0)
                except Exception:
                    pass
            # 该商户的优惠券
            item['coupons'] = [c for c in DEMO_COUPONS if c.get('merchant_id') == merchant_id and c.get('status', 1) == 1]
            return success(item)

        db = get_db()
        m = db.query_one("SELECT * FROM merchants WHERE id = %s AND status = 1", (merchant_id,))
        if not m:
            return success(None)
        if m.get('created_at') and hasattr(m['created_at'], 'strftime'):
            m['created_at'] = m['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        # 优惠券
        m['coupons'] = db.query_all(
            "SELECT * FROM coupon_templates WHERE merchant_id = %s AND status = 1",
            (merchant_id,))
        # purchase_links
        purchase_config = os.path.join(BASE_DIR, 'config', 'purchase_links.yaml')
        if os.path.exists(purchase_config):
            try:
                import yaml as _yaml
                with open(purchase_config, 'r', encoding='utf-8') as f:
                    plinks = _yaml.safe_load(f).get('purchase_links', {})
                pl = plinks.get(m['name'], {})
                if pl:
                    m['promo'] = pl.get('description', '')
                    m['purchase_link'] = pl.get('link', '')
                    m['purchase_name'] = pl.get('name', '')
                    m['purchase_price'] = pl.get('price', 0)
            except Exception:
                pass
        return success(m)
    except Exception as e:
        traceback.print_exc()
        return success(None)


@admin_bp.route('/h5/api/search', methods=['GET'])
def h5_search():
    """游客端 - 全局搜索（商户+活动+优惠券）"""
    try:
        keyword = request.args.get('keyword', '').strip()
        if not keyword:
            return success({'merchants': [], 'activities': [], 'coupons': []})

        if DEMO_MODE:
            merchants = [m for m in DEMO_MERCHANTS
                         if keyword in m['name'] or keyword in (m.get('description', '') or '') or keyword in (m.get('category', '') or '')]
            activities = [a for a in DEMO_ACTIVITIES
                          if keyword in a['title'] or keyword in (a.get('description', '') or '')]
            coupons = [c for c in DEMO_COUPONS
                       if keyword in c['name'] or keyword in (c.get('description', '') or '')]
            return success({'merchants': merchants[:10], 'activities': activities[:5], 'coupons': coupons[:5]})

        db = get_db()
        like = f'%{keyword}%'
        merchants = db.query_all(
            "SELECT id, name, category, floor, description FROM merchants WHERE status=1 AND (name LIKE %s OR description LIKE %s OR category LIKE %s) LIMIT 10",
            (like, like, like))
        activities = db.query_all(
            "SELECT id, title, start_at, end_at, location FROM activities WHERE status IN ('active','upcoming') AND (title LIKE %s OR description LIKE %s) LIMIT 5",
            (like, like))
        coupons = db.query_all(
            "SELECT t.id, t.name, t.type, t.value, t.min_amount, m.name as merchant_name FROM coupon_templates t LEFT JOIN merchants m ON t.merchant_id=m.id WHERE t.status=1 AND (t.name LIKE %s OR t.description LIKE %s) LIMIT 5",
            (like, like))
        return success({'merchants': merchants, 'activities': activities, 'coupons': coupons})
    except Exception as e:
        traceback.print_exc()
        return success({'merchants': [], 'activities': [], 'coupons': []})


# =============================================
# 数据导出 API
# =============================================

@admin_bp.route('/api/members/export', methods=['GET'])
@login_required
def export_members():
    """导出会员列表为Excel/CSV"""
    try:
        items = []
        if DEMO_MODE:
            items = DEMO_MEMBERS
        else:
            db = get_db()
            items = db.query_all("SELECT * FROM members ORDER BY id DESC")

        # 生成CSV
        import csv
        import io as _io
        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', '昵称', '手机号', '来源', '等级', '标签', '入园次数', '备注', '注册时间'])
        for m in items:
            tags = m.get('tags', '[]')
            if isinstance(tags, str):
                tags = json.loads(tags) if tags else []
            writer.writerow([
                m.get('id', ''),
                m.get('nickname', ''),
                m.get('phone', ''),
                m.get('source', ''),
                m.get('level', 'normal'),
                ', '.join(tags) if isinstance(tags, list) else str(tags),
                m.get('visit_count', 0),
                m.get('note', ''),
                m.get('created_at', '')
            ])

        from flask import make_response
        buf = _io.BytesIO()
        buf.write(output.getvalue().encode('utf-8-sig'))  # BOM for Excel
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = 'attachment; filename=members.csv'
        return resp
    except Exception as e:
        traceback.print_exc()
        return fail(f'导出失败: {str(e)}', 500)


@admin_bp.route('/api/merchants/export', methods=['GET'])
@login_required
def export_merchants():
    """导出商户列表为CSV"""
    try:
        items = []
        if DEMO_MODE:
            items = DEMO_MERCHANTS
        else:
            db = get_db()
            items = db.query_all("SELECT * FROM merchants ORDER BY id DESC")

        import csv
        import io as _io
        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ID', '商户名称', '分类', '位置', '楼层', '营业时间', '联系电话', '简介', '纬度', '经度', '购买链接', '状态', '创建时间'])
        for m in items:
            writer.writerow([
                m.get('id', ''),
                m.get('name', ''),
                m.get('category', ''),
                m.get('location', ''),
                m.get('floor', ''),
                m.get('open_hours', ''),
                m.get('phone', ''),
                m.get('description', ''),
                m.get('lat', ''),
                m.get('lng', ''),
                m.get('purchase_url', ''),
                '正常' if m.get('status', 1) == 1 else '下架',
                m.get('created_at', '')
            ])

        from flask import make_response
        buf = _io.BytesIO()
        buf.write(output.getvalue().encode('utf-8-sig'))
        buf.seek(0)
        resp = make_response(buf.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = 'attachment; filename=merchants.csv'
        return resp
    except Exception as e:
        traceback.print_exc()
        return fail(f'导出失败: {str(e)}', 500)


# =============================================
# 商户/活动批量导入 API
# =============================================

@admin_bp.route('/api/merchants/import', methods=['POST'])
@login_required
def import_merchants():
    """Excel批量导入商户"""
    try:
        if 'file' not in request.files:
            return fail('请上传文件')
        file = request.files['file']
        if not file.filename:
            return fail('文件为空')

        import io as _io
        try:
            import openpyxl
        except ImportError:
            return fail('服务器缺少Excel解析库，请执行：pip install openpyxl')

        file_bytes = file.read()
        wb = None
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return fail('无法解析文件，请确认为xlsx格式')

        if not rows or len(rows) < 2:
            return fail('文件为空或只有表头')

        # 解析表头
        header = [str(h or '').strip() for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if '名称' in h or '商户' in h:
                col_map['name'] = i
            elif '分类' in h or '类别' in h:
                col_map['category'] = i
            elif '位置' in h or '地址' in h:
                col_map['location'] = i
            elif '楼层' in h:
                col_map['floor'] = i
            elif '营业' in h or '时间' in h:
                col_map['open_hours'] = i
            elif '电话' in h or '联系' in h or '手机' in h:
                col_map['phone'] = i
            elif '简介' in h or '描述' in h or '说明' in h:
                col_map['description'] = i
            elif '纬度' in h or 'lat' in h.lower():
                col_map['lat'] = i
            elif '经度' in h or 'lng' in h.lower():
                col_map['lng'] = i
            elif '链接' in h or '购买' in h or 'url' in h.lower():
                col_map['purchase_url'] = i

        if 'name' not in col_map:
            return fail('表头缺少"商户名称"列')

        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue

                name = str(row[col_map['name']] or '').strip()
                if not name:
                    skipped += 1
                    continue

                category = str(row[col_map.get('category', -1)] or '').strip() if 'category' in col_map else '餐饮'
                location = str(row[col_map.get('location', -1)] or '').strip() if 'location' in col_map else ''
                floor = str(row[col_map.get('floor', -1)] or '').strip() if 'floor' in col_map else ''
                open_hours = str(row[col_map.get('open_hours', -1)] or '').strip() if 'open_hours' in col_map else ''
                phone = str(row[col_map.get('phone', -1)] or '').strip() if 'phone' in col_map else ''
                description = str(row[col_map.get('description', -1)] or '').strip() if 'description' in col_map else ''
                lat = row[col_map.get('lat', -1)] if 'lat' in col_map else None
                lng = row[col_map.get('lng', -1)] if 'lng' in col_map else None
                purchase_url = str(row[col_map.get('purchase_url', -1)] or '').strip() if 'purchase_url' in col_map else ''

                if DEMO_MODE:
                    if any(m['name'] == name for m in DEMO_MERCHANTS):
                        skipped += 1
                        continue
                    new_id = max(m['id'] for m in DEMO_MERCHANTS) + 1 if DEMO_MERCHANTS else 1
                    DEMO_MERCHANTS.append({
                        'id': new_id, 'name': name, 'category': category,
                        'location': location, 'floor': floor, 'open_hours': open_hours,
                        'phone': phone, 'description': description,
                        'lat': float(lat) if lat else None, 'lng': float(lng) if lng else None,
                        'purchase_url': purchase_url, 'status': 1,
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    _save_demo_data()
                else:
                    db = get_db()
                    existing = db.query_one("SELECT id FROM merchants WHERE name = %s", (name,))
                    if existing:
                        skipped += 1
                        continue
                    db.execute(
                        """INSERT INTO merchants (name, category, location, floor, open_hours, phone, description, lat, lng, purchase_url)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                        (name, category, location, floor, open_hours, phone, description, lat, lng, purchase_url or None)
                    )
                imported += 1
            except Exception as e:
                errors.append(f'第{row_idx}行：{str(e)}')
                skipped += 1

        result = {'imported': imported, 'skipped': skipped, 'errors': errors[:10]}
        return success(result, f'导入完成：成功{imported}条，跳过{skipped}条')
    except Exception as e:
        traceback.print_exc()
        return fail(f'导入失败: {str(e)}', 500)


@admin_bp.route('/api/activities/import', methods=['POST'])
@login_required
def import_activities():
    """Excel批量导入活动"""
    try:
        if 'file' not in request.files:
            return fail('请上传文件')
        file = request.files['file']
        if not file.filename:
            return fail('文件为空')

        import io as _io
        try:
            import openpyxl
        except ImportError:
            return fail('服务器缺少Excel解析库，请执行：pip install openpyxl')

        file_bytes = file.read()
        wb = None
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return fail('无法解析文件，请确认为xlsx格式')

        if not rows or len(rows) < 2:
            return fail('文件为空或只有表头')

        # 解析表头
        header = [str(h or '').strip() for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if '标题' in h or '活动' in h or '名称' in h:
                col_map['title'] = i
            elif '副标题' in h:
                col_map['subtitle'] = i
            elif '详情' in h or '内容' in h or '描述' in h:
                col_map['content'] = i
            elif '地点' in h or '位置' in h:
                col_map['location'] = i
            elif '开始' in h or '起始' in h:
                col_map['start_at'] = i
            elif '结束' in h or '截止' in h:
                col_map['end_at'] = i
            elif '人数' in h or '限制' in h:
                col_map['max_people'] = i
            elif '状态' in h:
                col_map['status'] = i

        if 'title' not in col_map:
            return fail('表头缺少"标题"列')

        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                if not row or all(c is None or str(c).strip() == '' for c in row):
                    continue

                title = str(row[col_map['title']] or '').strip()
                if not title:
                    skipped += 1
                    continue

                subtitle = str(row[col_map.get('subtitle', -1)] or '').strip() if 'subtitle' in col_map else ''
                content = str(row[col_map.get('content', -1)] or '').strip() if 'content' in col_map else ''
                location = str(row[col_map.get('location', -1)] or '').strip() if 'location' in col_map else ''
                max_people_raw = row[col_map.get('max_people', -1)] if 'max_people' in col_map else 0
                max_people = int(max_people_raw) if max_people_raw else 0
                status_raw = str(row[col_map.get('status', -1)] or '').strip() if 'status' in col_map else 'draft'
                status = status_raw if status_raw in ('draft', 'active', 'ended') else 'draft'

                # 解析时间
                def parse_dt(v):
                    if v is None:
                        return None
                    if isinstance(v, datetime):
                        return v
                    if isinstance(v, str):
                        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d', '%Y/%m/%d'):
                            try:
                                return datetime.strptime(v.strip(), fmt)
                            except ValueError:
                                continue
                    return None

                start_at = parse_dt(row[col_map.get('start_at', -1)]) if 'start_at' in col_map else datetime.now()
                end_at = parse_dt(row[col_map.get('end_at', -1)]) if 'end_at' in col_map else datetime.now() + timedelta(days=7)
                if not start_at:
                    start_at = datetime.now()
                if not end_at:
                    end_at = datetime.now() + timedelta(days=7)

                if DEMO_MODE:
                    if any(a['title'] == title for a in DEMO_ACTIVITIES):
                        skipped += 1
                        continue
                    new_id = max(a['id'] for a in DEMO_ACTIVITIES) + 1 if DEMO_ACTIVITIES else 1
                    DEMO_ACTIVITIES.append({
                        'id': new_id, 'title': title, 'subtitle': subtitle,
                        'content': content, 'location': location,
                        'start_at': start_at.strftime('%Y-%m-%d %H:%M:%S'),
                        'end_at': end_at.strftime('%Y-%m-%d %H:%M:%S'),
                        'max_people': max_people, 'signup_count': 0,
                        'status': status, 'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    })
                    _save_demo_data()
                else:
                    db = get_db()
                    existing = db.query_one("SELECT id FROM activities WHERE title = %s", (title,))
                    if existing:
                        skipped += 1
                        continue
                    db.execute(
                        """INSERT INTO activities (title, subtitle, content, location, start_at, end_at, max_people, status)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                        (title, subtitle, content, location, start_at, end_at, max_people, status)
                    )
                imported += 1
            except Exception as e:
                errors.append(f'第{row_idx}行：{str(e)}')
                skipped += 1

        result = {'imported': imported, 'skipped': skipped, 'errors': errors[:10]}
        return success(result, f'导入完成：成功{imported}条，跳过{skipped}条')
    except Exception as e:
        traceback.print_exc()
        return fail(f'导入失败: {str(e)}', 500)


# =============================================
# 独立运行（调试用）：把 Blueprint 挂到独立 app 上
# =============================================

app.register_blueprint(admin_bp)


if __name__ == '__main__':
    import io
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    mode_str = "[演示模式 - 无需数据库]" if DEMO_MODE else "[正式模式 - 已连接数据库]"
    print("=" * 55)
    print(f"  园区AI助手 - 管理后台 API 服务  {mode_str}")
    print(f"  访问地址: http://localhost:8082/admin")
    print(f"  商户核销: http://localhost:8082/merchant-verify")
    print(f"  API 地址: http://localhost:8082/api")
    print(f"  默认账号: admin / admin123")
    print("=" * 55)
    app.run(host='0.0.0.0', port=8082, debug=False)
